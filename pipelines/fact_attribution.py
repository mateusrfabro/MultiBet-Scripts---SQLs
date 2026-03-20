"""
Pipeline: fact_attribution (Consolidacao de Atribuicao)
=======================================================
Dominio 2 — Aquisicao de Jogadores (Prioridade 2)

Consolida:
    1. fact_registrations           -> qty_registrations por dt
    2. fact_ftd_deposits            -> qty_ftds por dt
    3. fact_gaming_activity_daily   -> ggr por dt (agregado de todos trackers)
    4. Google Sheets (Analise Total Google) -> marketing_spend por dt

Dimensao: dt (dia)
Metricas brutas: qty_registrations, qty_ftds, ggr, marketing_spend
Metricas calculadas (na View): CPA, CAC, ROAS, ROI

Chave de atribuicao: Google Ads spend atribuido ao total do dia (sem quebra por tracker).

Destino: Super Nova DB -> multibet.fact_attribution
Estrategia: TRUNCATE + INSERT

Execucao:
    python pipelines/fact_attribution.py
"""

import sys
import os
import logging
from datetime import datetime, timezone

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from db.supernova import execute_supernova, get_supernova_connection

import pandas as pd
import openpyxl
import psycopg2.extras

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# Caminho da planilha Google Ads (exportada do Google Sheets)
GOOGLE_SHEETS_PATH = "C:/Users/NITRO/Downloads/Análise Total Google.xlsx"

# --- DDL ----------------------------------------------------------------------
DDL_SCHEMA = "CREATE SCHEMA IF NOT EXISTS multibet;"

DDL_TABLE = """
CREATE TABLE IF NOT EXISTS multibet.fact_attribution (
    dt                      DATE,
    c_tracker_id            VARCHAR(255),
    qty_registrations       INTEGER DEFAULT 0,
    qty_ftds                INTEGER DEFAULT 0,
    ggr                     NUMERIC(18,2) DEFAULT 0,
    marketing_spend         NUMERIC(18,2) DEFAULT 0,
    refreshed_at            TIMESTAMPTZ DEFAULT NOW(),
    PRIMARY KEY (dt, c_tracker_id)
);
"""

# View para calculos derivados (CPA, CAC, ROAS)
DDL_DROP_VIEW = "DROP VIEW IF EXISTS multibet.vw_attribution_metrics;"
DDL_VIEW = """
CREATE VIEW multibet.vw_attribution_metrics AS
SELECT
    dt,
    c_tracker_id,
    qty_registrations,
    qty_ftds,
    ggr,
    marketing_spend,
    CASE WHEN qty_ftds > 0
         THEN ROUND(marketing_spend / qty_ftds, 2)
         ELSE NULL END AS cpa,
    CASE WHEN qty_registrations > 0
         THEN ROUND(marketing_spend / qty_registrations, 2)
         ELSE NULL END AS cac,
    CASE WHEN marketing_spend > 0
         THEN ROUND(ggr / marketing_spend, 4)
         ELSE NULL END AS roas,
    CASE WHEN marketing_spend > 0
         THEN ROUND((ggr - marketing_spend) / marketing_spend * 100, 2)
         ELSE NULL END AS roi_pct,
    refreshed_at
FROM multibet.fact_attribution;
"""


def load_google_spend() -> pd.DataFrame:
    """Le todas as abas da planilha Google Ads e consolida spend diario."""
    log.info(f"Lendo planilha Google Ads: {GOOGLE_SHEETS_PATH}")

    wb = openpyxl.load_workbook(GOOGLE_SHEETS_PATH, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log.info(f"  Aba: {sheet_name} ({ws.max_row} rows)")

        for row in range(2, ws.max_row + 1):
            dt_val = ws.cell(row=row, column=1).value   # A: data
            spend_val = ws.cell(row=row, column=2).value  # B: INVESTIDO

            # Validar data
            if not isinstance(dt_val, datetime):
                continue
            # Validar spend
            if spend_val is None or not isinstance(spend_val, (int, float)):
                continue
            if spend_val <= 0:
                continue

            all_rows.append({
                "dt": dt_val.date(),
                "c_tracker_id": "google_ads",
                "marketing_spend": round(float(spend_val), 2),
            })

    df = pd.DataFrame(all_rows)
    if df.empty:
        return df

    # Agregar por dia + tracker (caso haja duplicatas entre abas)
    df = df.groupby(["dt", "c_tracker_id"], as_index=False).agg({"marketing_spend": "sum"})
    df = df.sort_values("dt")

    log.info(f"  {len(df)} dias com spend. Total: R$ {df['marketing_spend'].sum():,.2f}")
    return df


def load_fact_tables() -> pd.DataFrame:
    """Le as 3 fact tables do Super Nova DB por dt + c_tracker_id."""
    log.info("Lendo fact tables do Super Nova DB (por tracker)...")

    # Registros por dia + tracker (fact_registrations nao tem tracker,
    # entao puxamos direto do Athena via uma query no Super Nova DB
    # que cruza fact_registrations com a contagem por tracker da gaming)
    # Alternativa: usar fact_gaming_activity_daily que ja tem tracker
    # Para registros por tracker, usamos a bireports via Athena
    from db.athena import query_athena

    log.info("  Buscando registros por tracker no Athena...")
    df_reg = query_athena("""
        SELECT
            CAST(c_sign_up_time AT TIME ZONE 'UTC' AT TIME ZONE 'America/Sao_Paulo' AS DATE) AS dt,
            COALESCE(NULLIF(TRIM(c_tracker_id), ''), CAST(c_affiliate_id AS VARCHAR), 'sem_tracker') AS c_tracker_id,
            COUNT(DISTINCT c_ecr_id) AS qty_registrations
        FROM bireports_ec2.tbl_ecr
        WHERE c_sign_up_time >= TIMESTAMP '2025-10-01'
        GROUP BY 1, 2
    """, database="bireports_ec2")
    log.info(f"  Registros por tracker: {len(df_reg)} linhas")

    # FTDs por dia + tracker (agora tem tracker na V3)
    rows_ftd = execute_supernova(
        "SELECT dt, c_tracker_id, qty_ftds FROM multibet.fact_ftd_deposits ORDER BY dt",
        fetch=True,
    )
    df_ftd = pd.DataFrame(rows_ftd, columns=["dt", "c_tracker_id", "qty_ftds"])
    log.info(f"  fact_ftd_deposits: {len(df_ftd)} linhas")

    # GGR por dia + tracker
    rows_ggr = execute_supernova(
        "SELECT dt, c_tracker_id, ggr FROM multibet.fact_gaming_activity_daily ORDER BY dt",
        fetch=True,
    )
    df_ggr = pd.DataFrame(rows_ggr, columns=["dt", "c_tracker_id", "ggr"])
    log.info(f"  fact_gaming_activity_daily: {len(df_ggr)} linhas")

    # FULL OUTER JOIN por dt + c_tracker_id
    df = df_reg.merge(df_ftd, on=["dt", "c_tracker_id"], how="outer")
    df = df.merge(df_ggr, on=["dt", "c_tracker_id"], how="outer")
    df = df.fillna(0)

    return df


def setup_table():
    log.info("Verificando/criando tabela e view multibet.fact_attribution...")
    execute_supernova(DDL_SCHEMA)
    execute_supernova(DDL_TABLE)
    execute_supernova(DDL_DROP_VIEW)
    execute_supernova(DDL_VIEW)
    log.info("Tabela e view prontas.")


def load_mapping() -> dict:
    """Carrega dim_marketing_mapping: tracker_id -> source."""
    rows = execute_supernova(
        "SELECT tracker_id, source FROM multibet.dim_marketing_mapping",
        fetch=True,
    )
    mapping = {r[0]: r[1] for r in (rows or [])}
    log.info(f"  dim_marketing_mapping: {len(mapping)} trackers mapeados")
    return mapping


def refresh():
    # 1. Carregar spend do Google Sheets (total diario)
    df_spend = load_google_spend()

    # 2. Carregar fact tables (por dt + tracker)
    df_facts = load_fact_tables()

    # 3. Carregar mapeamento tracker -> source
    mapping = load_mapping()
    google_trackers = [t for t, s in mapping.items() if s == "google_ads"]
    log.info(f"  Trackers Google Ads: {google_trackers}")

    # 4. Distribuir spend proporcionalmente por FTDs dos trackers Google
    df_facts["source"] = df_facts["c_tracker_id"].map(mapping).fillna("outros")
    df_facts["marketing_spend"] = 0.0

    if not df_spend.empty and google_trackers:
        # Para cada dia, distribuir spend proporcionalmente aos FTDs dos trackers Google
        spend_by_date = df_spend.set_index("dt")["marketing_spend"].to_dict()

        for dt_val, spend_val in spend_by_date.items():
            # FTDs dos trackers Google neste dia
            mask = (df_facts["dt"] == dt_val) & (df_facts["c_tracker_id"].isin(google_trackers))
            ftds_google = df_facts.loc[mask, "qty_ftds"]
            total_ftds_day = ftds_google.sum()

            if total_ftds_day > 0:
                # Distribuir proporcionalmente
                df_facts.loc[mask, "marketing_spend"] = (
                    ftds_google / total_ftds_day * spend_val
                )
            else:
                # Sem FTDs: distribuir igualmente entre trackers Google ativos
                active = df_facts.loc[mask]
                if len(active) > 0:
                    df_facts.loc[mask, "marketing_spend"] = spend_val / len(active)
                else:
                    # Nenhum tracker Google ativo: criar linha 'google_ads' para nao perder spend
                    new_row = pd.DataFrame([{
                        "dt": dt_val, "c_tracker_id": "google_ads",
                        "qty_registrations": 0, "qty_ftds": 0, "ggr": 0,
                        "source": "google_ads", "marketing_spend": spend_val
                    }])
                    df_facts = pd.concat([df_facts, new_row], ignore_index=True)

    df = df_facts.drop(columns=["source"])
    df = df.fillna(0)
    df = df.sort_values("dt")

    log.info(f"{len(df)} dias totais para inserir.")

    # 4. Inserir no Super Nova DB
    now_utc = datetime.now(timezone.utc)

    insert_sql = """
        INSERT INTO multibet.fact_attribution
            (dt, c_tracker_id, qty_registrations, qty_ftds, ggr, marketing_spend, refreshed_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """

    records = []
    for _, row in df.iterrows():
        tracker = row["c_tracker_id"] if row["c_tracker_id"] else "sem_tracker"
        records.append((
            row["dt"],
            tracker,
            int(row["qty_registrations"]),
            int(row["qty_ftds"]),
            float(row["ggr"]),
            float(row["marketing_spend"]),
            now_utc,
        ))

    ssh, conn = get_supernova_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("TRUNCATE TABLE multibet.fact_attribution;")
            psycopg2.extras.execute_batch(cur, insert_sql, records, page_size=500)
        conn.commit()
    finally:
        conn.close()
        ssh.close()

    total_spend = sum(r[5] for r in records)
    total_ftds = sum(r[3] for r in records)
    total_ggr = sum(r[4] for r in records)
    cpa_global = total_spend / max(total_ftds, 1)
    roas_global = total_ggr / max(total_spend, 1)

    log.info(f"{len(records)} dias inseridos | "
             f"Spend: R$ {total_spend:,.2f} | FTDs: {total_ftds:,} | "
             f"GGR: R$ {total_ggr:,.2f} | "
             f"CPA global: R$ {cpa_global:,.2f} | ROAS: {roas_global:.2f}x")


if __name__ == "__main__":
    log.info("=== Iniciando pipeline fact_attribution ===")
    setup_table()
    refresh()
    log.info("=== Pipeline concluido ===")

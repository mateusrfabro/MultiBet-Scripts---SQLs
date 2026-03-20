"""
Gera Excel com o dicionário de dados das tabelas canônicas P2.
Formato para apresentação ao Head e equipe.

KPIs seguem EXATAMENTE o modelo definido no planejamento.

Execução:
    python pipelines/gerar_dicionario_dados.py

Saída: docs/dicionario_tabelas_canonicas_P2.xlsx
"""

import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Aba 1: Visão Geral — KPIs exatos do planejamento + 5 colunas de rastreabilidade ─
# Formato: Domínio | Tabela | Tipo | KPI / Indicador | Fórmula / Descrição | Prioridade |
#          Fórmula / Descrição - Dados | Fonte | Database Origem | Tabela Origem | Coluna Origem | Dados (exemplo)

visao_geral = pd.DataFrame([
    # ─── fact_registrations (Prioridade 2) ─────────────────────────────────────
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_registrations",
        "Tipo": "FACT",
        "KPI / Indicador": "Cadastros totais / dia",
        "Fórmula / Descrição": "Novos registros por dia",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "COUNT(*) por registration_date",
        "Fonte": "Athena (Pragmatic Solutions)",
        "Database Origem": "ecr_ec2",
        "Tabela Origem": "tbl_ecr",
        "Coluna Origem": "c_signup_time, c_ecr_id",
        "Dados (exemplo)": "2026-03-18: 1.247 cadastros",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_registrations",
        "Tipo": "FACT",
        "KPI / Indicador": "FTDs (First Time Depositors)",
        "Fórmula / Descrição": "Jogadores com 1° depósito",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "COUNT(DISTINCT c_ecr_id) FROM fact_ftd_deposits por ftd_date",
        "Fonte": "Athena (Pragmatic Solutions)",
        "Database Origem": "cashier_ec2",
        "Tabela Origem": "tbl_cashier_deposit",
        "Coluna Origem": "c_ecr_id, c_created_time (ROW_NUMBER rn=1)",
        "Dados (exemplo)": "2026-03-18: 842 FTDs",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_registrations",
        "Tipo": "FACT",
        "KPI / Indicador": "FTD Rate %",
        "Fórmula / Descrição": "FTDs ÷ Cadastros",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "COUNT(fact_ftd_deposits) / COUNT(fact_registrations) × 100",
        "Fonte": "Super Nova DB (calculado)",
        "Database Origem": "multibet",
        "Tabela Origem": "fact_registrations + fact_ftd_deposits",
        "Coluna Origem": "c_ecr_id (JOIN entre as duas tabelas)",
        "Dados (exemplo)": "67,5%",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_registrations",
        "Tipo": "FACT",
        "KPI / Indicador": "Tempo cadastro → FTD (h)",
        "Fórmula / Descrição": "Horas entre registro e 1° depósito",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "AVG(ftd_time - registration_time) em horas",
        "Fonte": "Super Nova DB (calculado)",
        "Database Origem": "multibet",
        "Tabela Origem": "fact_registrations JOIN fact_ftd_deposits ON c_ecr_id",
        "Coluna Origem": "registration_time, ftd_time",
        "Dados (exemplo)": "Média: 2,4h",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_registrations",
        "Tipo": "FACT",
        "KPI / Indicador": "Taxa verificação KYC",
        "Fórmula / Descrição": "KYCs aprovados ÷ Cadastros",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "COUNT(kyc_verified) / COUNT(registrations) × 100",
        "Fonte": "Athena (Pragmatic Solutions)",
        "Database Origem": "ecr_ec2",
        "Tabela Origem": "tbl_ecr (campo KYC a validar) ou csm_ec2",
        "Coluna Origem": "⚠️ Pendente validação — verificar se campo KYC existe em ecr_ec2 ou csm_ec2",
        "Dados (exemplo)": "Dado ainda não disponível — validar fonte",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_registrations",
        "Tipo": "FACT",
        "KPI / Indicador": "Cadastros por dispositivo",
        "Fórmula / Descrição": "Mobile vs Desktop",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "COUNT(*) GROUP BY canal/dispositivo",
        "Fonte": "Athena (Pragmatic Solutions)",
        "Database Origem": "ecr_ec2 ou fund_ec2",
        "Tabela Origem": "tbl_ecr (campo device a validar) ou tbl_real_fund_txn.c_channel",
        "Coluna Origem": "⚠️ Pendente validação — tbl_ecr não tem c_channel; fund_ec2 tem c_channel (DESKTOP/MOBILE)",
        "Dados (exemplo)": "Dado pendente — validar se ecr_ec2 tem info de device",
    },

    # ─── fact_ftd_deposits (Prioridade 2) ──────────────────────────────────────
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_ftd_deposits",
        "Tipo": "FACT",
        "KPI / Indicador": "Ticket médio FTD",
        "Fórmula / Descrição": "Valor médio do 1° depósito",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "AVG(ftd_amount) ou SUM(ftd_amount) / COUNT(DISTINCT c_ecr_id)",
        "Fonte": "Athena (Pragmatic Solutions)",
        "Database Origem": "cashier_ec2",
        "Tabela Origem": "tbl_cashier_deposit",
        "Coluna Origem": "c_confirmed_amount_in_inhouse_ccy (÷100 → BRL)",
        "Dados (exemplo)": "R$ 47,20",
    },

    # ─── fact_attribution (Prioridade 2) ───────────────────────────────────────
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_attribution",
        "Tipo": "FACT",
        "KPI / Indicador": "CPA (Custo por Aquisição)",
        "Fórmula / Descrição": "Spend ÷ FTDs",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "SUM(spend) / COUNT(DISTINCT ftd_ecr_id)",
        "Fonte": "⚠️ Spend vem de fonte EXTERNA (Google/Meta Ads)",
        "Database Origem": "Fonte externa + multibet",
        "Tabela Origem": "Tabela de spend (a criar) + fact_ftd_deposits",
        "Coluna Origem": "⚠️ Spend não disponível no Athena nem Smartico — precisa integração com plataformas de mídia",
        "Dados (exemplo)": "Dado pendente — requer fonte de spend",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_attribution",
        "Tipo": "FACT",
        "KPI / Indicador": "CAC (Custo Aquisição Cliente)",
        "Fórmula / Descrição": "Custo total ÷ Novos clientes",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "SUM(custo_total) / COUNT(DISTINCT new_ecr_id)",
        "Fonte": "⚠️ Custo vem de fonte EXTERNA",
        "Database Origem": "Fonte externa + multibet",
        "Tabela Origem": "Tabela de custos (a criar) + fact_registrations",
        "Coluna Origem": "⚠️ Custo não disponível no Athena nem Smartico — precisa integração",
        "Dados (exemplo)": "Dado pendente — requer fonte de custo",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "fact_attribution",
        "Tipo": "FACT",
        "KPI / Indicador": "Receita por afiliado",
        "Fórmula / Descrição": "NGR atribuído por afiliado",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "SUM(ngr) GROUP BY c_tracker_id, via JOIN fact_registrations + ps_bi.fct_player_activity_daily",
        "Fonte": "Athena (ps_bi) + Super Nova DB",
        "Database Origem": "ps_bi + multibet",
        "Tabela Origem": "fct_player_activity_daily JOIN fact_registrations",
        "Coluna Origem": "ps_bi: ggr, btr, rca (NGR = GGR - BTR - RCA) | multibet: c_tracker_id",
        "Dados (exemplo)": "Tracker 297657: R$ 12.450 NGR",
    },

    # ─── dim_acquisition_channel (Prioridade 2) ───────────────────────────────
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "dim_acquisition_channel",
        "Tipo": "DIM",
        "KPI / Indicador": "Mix de canais %",
        "Fórmula / Descrição": "Participação de cada canal no total",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "COUNT(*) por canal / COUNT(*) total × 100",
        "Fonte": "Athena (Pragmatic Solutions)",
        "Database Origem": "ecr_ec2",
        "Tabela Origem": "tbl_ecr + tbl_ecr_banner (reference_url com UTMs)",
        "Coluna Origem": "c_tracker_id + c_reference_url (parsing gclid/fbclid/ttclid)",
        "Dados (exemplo)": "Orgânico: 45%, Google: 30%, Meta: 15%, Afiliados: 10%",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "dim_acquisition_channel",
        "Tipo": "DIM",
        "KPI / Indicador": "Qualidade de tráfego por canal",
        "Fórmula / Descrição": "FTD Rate e LTV por canal",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "FTD Rate % e AVG(LTV) GROUP BY canal",
        "Fonte": "Super Nova DB (calculado) + Athena (ps_bi)",
        "Database Origem": "multibet + ps_bi",
        "Tabela Origem": "dim_acquisition_channel JOIN fact_ftd + fct_player_activity_daily",
        "Coluna Origem": "canal, ftd_rate, ltv (calculados)",
        "Dados (exemplo)": "Google: FTD Rate 72%, LTV R$180 | Meta: FTD Rate 58%, LTV R$120",
    },

    # ─── agg_cohort_acquisition (Prioridade 2) ────────────────────────────────
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "agg_cohort_acquisition",
        "Tipo": "AGG",
        "KPI / Indicador": "LTV por cohort (D7/D30/D90)",
        "Fórmula / Descrição": "LTV acumulado por coorte de FTD",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "SUM(deposits - withdrawals) por cohort_date nos intervalos D+7, D+30, D+90",
        "Fonte": "Athena (ps_bi)",
        "Database Origem": "ps_bi + multibet",
        "Tabela Origem": "fct_player_activity_daily JOIN fact_ftd_deposits (cohort por ftd_date)",
        "Coluna Origem": "deposits, withdrawals, activity_date, ftd_date (cohort anchor)",
        "Dados (exemplo)": "Cohort Mar/26: D7=R$85, D30=R$210, D90=R$450",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "agg_cohort_acquisition",
        "Tipo": "AGG",
        "KPI / Indicador": "2nd deposit rate",
        "Fórmula / Descrição": "Jogadores com 2° depósito ÷ FTDs",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "COUNT(players com ≥2 depósitos) / COUNT(FTDs) × 100",
        "Fonte": "Athena (Pragmatic Solutions)",
        "Database Origem": "cashier_ec2 + multibet",
        "Tabela Origem": "tbl_cashier_deposit (ROW_NUMBER rn=2) JOIN fact_ftd_deposits",
        "Coluna Origem": "c_ecr_id, c_created_time, c_txn_status = 'txn_confirmed_success'",
        "Dados (exemplo)": "38% dos FTDs fazem 2° depósito",
    },
    {
        "Domínio": "Aquisição de Jogadores",
        "Tabela": "agg_cohort_acquisition",
        "Tipo": "AGG",
        "KPI / Indicador": "GGR acumulado por cohort",
        "Fórmula / Descrição": "GGR total desde FTD por coorte",
        "Prioridade": 2,
        "Fórmula / Descrição - Dados": "SUM(ggr) por cohort_date acumulado desde ftd_date",
        "Fonte": "Athena (ps_bi)",
        "Database Origem": "ps_bi + multibet",
        "Tabela Origem": "fct_player_activity_daily JOIN fact_ftd_deposits (cohort por ftd_date)",
        "Coluna Origem": "ggr, activity_date, ftd_date",
        "Dados (exemplo)": "Cohort Fev/26: GGR acumulado = R$ 2.1M",
    },
])

# ─── Aba 2: Dicionário de Colunas (tabelas já criadas) ───────────────────────
dicionario_colunas = pd.DataFrame([
    # fact_registrations
    {"Tabela": "fact_registrations", "Coluna": "c_ecr_id", "Tipo PG": "BIGINT", "Fonte": "Athena ecr_ec2.tbl_ecr", "Coluna Origem": "c_ecr_id", "Transformação": "Nenhuma", "Descrição": "ID interno do jogador (PK lógica, unique)"},
    {"Tabela": "fact_registrations", "Coluna": "c_external_id", "Tipo PG": "BIGINT", "Fonte": "Athena ecr_ec2.tbl_ecr", "Coluna Origem": "c_external_id", "Transformação": "CAST BIGINT", "Descrição": "ID externo (= Smartico user_ext_id)"},
    {"Tabela": "fact_registrations", "Coluna": "c_tracker_id", "Tipo PG": "VARCHAR(255)", "Fonte": "Athena ecr_ec2.tbl_ecr", "Coluna Origem": "c_tracker_id", "Transformação": "Nenhuma", "Descrição": "ID do tracker/afiliado de origem"},
    {"Tabela": "fact_registrations", "Coluna": "c_country_code", "Tipo PG": "VARCHAR(50)", "Fonte": "Athena ecr_ec2.tbl_ecr", "Coluna Origem": "c_jurisdiction", "Transformação": "Renomeado", "Descrição": "País do jogador (ex: brazil)"},
    {"Tabela": "fact_registrations", "Coluna": "registration_date", "Tipo PG": "DATE", "Fonte": "Athena ecr_ec2.tbl_ecr", "Coluna Origem": "c_signup_time", "Transformação": "UTC→BRT + CAST DATE", "Descrição": "Data do registro em BRT"},
    {"Tabela": "fact_registrations", "Coluna": "registration_time", "Tipo PG": "TIMESTAMPTZ", "Fonte": "Athena ecr_ec2.tbl_ecr", "Coluna Origem": "c_signup_time", "Transformação": "UTC original", "Descrição": "Timestamp exato do registro"},
    {"Tabela": "fact_registrations", "Coluna": "dt", "Tipo PG": "DATE", "Fonte": "Derivada", "Coluna Origem": "= registration_date", "Transformação": "Cópia", "Descrição": "Partição lógica"},
    # fact_ftd_deposits
    {"Tabela": "fact_ftd_deposits", "Coluna": "c_ecr_id", "Tipo PG": "BIGINT", "Fonte": "Athena cashier_ec2", "Coluna Origem": "c_ecr_id", "Transformação": "Nenhuma", "Descrição": "ID interno do jogador (PK lógica, unique)"},
    {"Tabela": "fact_ftd_deposits", "Coluna": "ftd_txn_id", "Tipo PG": "BIGINT", "Fonte": "Athena cashier_ec2", "Coluna Origem": "c_txn_id", "Transformação": "Renomeado", "Descrição": "ID da transação do primeiro depósito"},
    {"Tabela": "fact_ftd_deposits", "Coluna": "ftd_amount", "Tipo PG": "NUMERIC(15,2)", "Fonte": "Athena cashier_ec2", "Coluna Origem": "c_confirmed_amount_in_inhouse_ccy", "Transformação": "÷ 100 (centavos → BRL)", "Descrição": "Valor do 1° depósito em R$"},
    {"Tabela": "fact_ftd_deposits", "Coluna": "ftd_date", "Tipo PG": "DATE", "Fonte": "Athena cashier_ec2", "Coluna Origem": "c_created_time", "Transformação": "UTC→BRT + CAST DATE", "Descrição": "Data do FTD em BRT"},
    {"Tabela": "fact_ftd_deposits", "Coluna": "ftd_time", "Tipo PG": "TIMESTAMPTZ", "Fonte": "Athena cashier_ec2", "Coluna Origem": "c_created_time", "Transformação": "UTC original", "Descrição": "Timestamp exato do FTD"},
    {"Tabela": "fact_ftd_deposits", "Coluna": "dt", "Tipo PG": "DATE", "Fonte": "Derivada", "Coluna Origem": "= ftd_date", "Transformação": "Cópia", "Descrição": "Partição lógica"},
])

# ─── Aba 3: Regras de Negócio ────────────────────────────────────────────────
regras_negocio = pd.DataFrame([
    {"Tabela": "fact_registrations", "Regra": "Filtro de status", "Detalhe": "Apenas c_registration_status = 'ecr_regn_completed'"},
    {"Tabela": "fact_registrations", "Regra": "Fuso horário", "Detalhe": "registration_date: UTC→BRT (AT TIME ZONE 'America/Sao_Paulo')"},
    {"Tabela": "fact_registrations", "Regra": "Unicidade", "Detalhe": "1 registro por c_ecr_id (UNIQUE INDEX, ON CONFLICT DO NOTHING)"},
    {"Tabela": "fact_registrations", "Regra": "Carga", "Detalhe": "Incremental por c_signup_time > último salvo"},
    {"Tabela": "fact_ftd_deposits", "Regra": "Definição FTD", "Detalhe": "ROW_NUMBER() PARTITION BY c_ecr_id ORDER BY c_created_time ASC, WHERE rn = 1"},
    {"Tabela": "fact_ftd_deposits", "Regra": "Filtro de status", "Detalhe": "c_txn_status = 'txn_confirmed_success' (NÃO 'SUCCESS' — validado 18/03)"},
    {"Tabela": "fact_ftd_deposits", "Regra": "Conversão de valor", "Detalhe": "c_confirmed_amount_in_inhouse_ccy em centavos → ÷100 para BRL"},
    {"Tabela": "fact_ftd_deposits", "Regra": "Unicidade", "Detalhe": "1 registro por c_ecr_id (cada jogador tem no máximo 1 FTD)"},
    {"Tabela": "fact_attribution", "Regra": "Dependência", "Detalhe": "⚠️ CPA e CAC dependem de SPEND de mídia (fonte externa — não disponível no Athena/Smartico)"},
    {"Tabela": "fact_attribution", "Regra": "NGR por afiliado", "Detalhe": "JOIN fact_registrations.c_tracker_id + ps_bi.fct_player_activity_daily (NGR = GGR - BTR - RCA)"},
    {"Tabela": "dim_acquisition_channel", "Regra": "Parsing UTMs", "Detalhe": "ecr_ec2.tbl_ecr_banner.c_reference_url → extrair gclid, fbclid, ttclid para classificar canal"},
    {"Tabela": "agg_cohort_acquisition", "Regra": "Cohort anchor", "Detalhe": "Coorte definida por ftd_date (mês do primeiro depósito). LTV/GGR acumulados desde D+0"},
])

# ─── Aba 4: Status de Implementação ──────────────────────────────────────────
status_impl = pd.DataFrame([
    {"Tabela": "fact_registrations", "Status": "✅ IMPLEMENTADA", "Registros": "359.859", "Pipeline": "pipelines/fact_registrations.py", "Observação": "Carga inicial completa. Incremental pronto."},
    {"Tabela": "fact_ftd_deposits", "Status": "✅ IMPLEMENTADA", "Registros": "257.569", "Pipeline": "pipelines/fact_ftd_deposits.py", "Observação": "Carga inicial completa. Incremental pronto."},
    {"Tabela": "agg_financial_monthly", "Status": "✅ VIEW CRIADA", "Registros": "—", "Pipeline": "pipelines/ddl_agg_financial_monthly.sql", "Observação": "View sobre fact_ftd_deposits."},
    {"Tabela": "fact_attribution", "Status": "⚠️ PARCIAL", "Registros": "—", "Pipeline": "A criar", "Observação": "NGR por afiliado viável via ps_bi. CPA/CAC dependem de fonte externa de spend."},
    {"Tabela": "dim_acquisition_channel", "Status": "⚠️ PARCIAL", "Registros": "—", "Pipeline": "A criar", "Observação": "c_tracker_id disponível. Parsing UTMs de tbl_ecr_banner viável (pipeline de_para existe)."},
    {"Tabela": "agg_cohort_acquisition", "Status": "🔜 PENDENTE", "Registros": "—", "Pipeline": "A criar", "Observação": "Depende de fact_ftd_deposits (✅) + ps_bi.fct_player_activity_daily. Query cohort a montar."},
])

# ─── Gerar Excel ──────────────────────────────────────────────────────────────
os.makedirs("docs", exist_ok=True)
output_path = "docs/dicionario_tabelas_canonicas_P2.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    visao_geral.to_excel(writer, sheet_name="Visão Geral", index=False)
    dicionario_colunas.to_excel(writer, sheet_name="Dicionário de Colunas", index=False)
    regras_negocio.to_excel(writer, sheet_name="Regras de Negócio", index=False)
    status_impl.to_excel(writer, sheet_name="Status Implementação", index=False)

    # ─── Formatação profissional ──────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ok_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]

        # Header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # Data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                # Highlight cells with warnings
                if cell.value and "⚠️" in str(cell.value):
                    cell.fill = warning_fill

        # Auto-width
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 55)

        # Freeze header
        ws.freeze_panes = "A2"
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

print(f"Excel gerado: {output_path}")

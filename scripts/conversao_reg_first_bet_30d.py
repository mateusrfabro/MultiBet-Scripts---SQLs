"""
Analise de Conversao Registro -> Primeira Aposta Gaming (Ultimos 30 dias)
=========================================================================
Objetivo: Medir impacto da mudanca no fluxo de cadastro (automatizacao
          de endereco) na conversao de registro para primeira aposta.

ATENCAO: "Primeira Aposta" != "Primeiro Deposito (FTD)".
  - First Bet = primeira transacao gaming de debito (c_op_type = 'DB',
    c_is_gaming_txn = 'Y', c_txn_status = 'SUCCESS')
  - FTD = primeiro deposito financeiro (tipo 1)

Fonte:
  - Registros: ecr_ec2.tbl_ecr (c_signup_time)
  - First Bets: fund_ec2.tbl_real_fund_txn (MIN c_start_time gaming DB)
  - Test users: ecr_ec2.tbl_ecr_flags (c_test_user = false)
  - Tipo txn: fund_ec2.tbl_real_fund_txn_type_mst (c_is_gaming_txn = 'Y')

Regras:
  - Timestamps convertidos UTC -> BRT (AT TIME ZONE)
  - Valores fund_ec2 em centavos (/100.0) -- porem nao usamos valores aqui
  - Sem SELECT * -- apenas colunas necessarias
  - CTEs (nunca CREATE TEMP TABLE)
  - Sintaxe Presto/Trino (Athena)
  - c_test_user = false (excluir usuarios de teste)

Perspectiva: COHORT por dia de registro
  - Para cada dia de registro, quantos jogadores ja fizeram first bet (ate hoje)?
  - Isso permite ver se a mudanca no cadastro impactou a conversao.

Data: 2026-03-24
Autor: Mateus F. (Squad Intelligence Engine)
"""

import sys
import os
import logging
from datetime import datetime, timedelta

# Garantir path do projeto
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Backend sem GUI (salvar PNG)
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils.dataframe import dataframe_to_rows

from db.athena import query_athena

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
log = logging.getLogger(__name__)

# ==============================================================================
# CONFIG
# ==============================================================================
# Periodo: ultimos 30 dias completos (D-31 a D-1, excluindo hoje que e parcial)
TODAY = datetime(2026, 3, 24)
DATE_END = (TODAY - timedelta(days=1)).strftime('%Y-%m-%d')    # 2026-03-23
DATE_START = (TODAY - timedelta(days=31)).strftime('%Y-%m-%d')  # 2026-02-21

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_PATH = os.path.join(OUTPUT_DIR, "conversao_reg_first_bet_30d.csv")
PNG_PATH = os.path.join(OUTPUT_DIR, "conversao_reg_first_bet_30d.png")
XLSX_PATH = os.path.join(OUTPUT_DIR, "conversao_reg_first_bet_30d.xlsx")


# ==============================================================================
# STEP 0: Validar schema (descoberta rapida)
# ==============================================================================
def discover_schemas():
    """Valida que as colunas esperadas existem antes da query pesada."""
    log.info("=== STEP 0: Validando schemas ===")

    # Verificar colunas de tbl_ecr
    sql_ecr = """
    SELECT column_name
    FROM information_schema.columns
    WHERE table_schema = 'ecr_ec2'
      AND table_name = 'tbl_ecr'
      AND column_name IN ('c_ecr_id', 'c_signup_time', 'c_external_id')
    """
    df_ecr = query_athena(sql_ecr, database="ecr_ec2")
    log.info(f"ecr_ec2.tbl_ecr colunas encontradas: {df_ecr['column_name'].tolist()}")

    # Verificar colunas de tbl_real_fund_txn_type_mst
    sql_mst = """
    SELECT column_name
    FROM information_schema.columns
    WHERE table_schema = 'fund_ec2'
      AND table_name = 'tbl_real_fund_txn_type_mst'
      AND column_name IN ('c_txn_type', 'c_is_gaming_txn', 'c_op_type')
    """
    df_mst = query_athena(sql_mst, database="fund_ec2")
    log.info(f"fund_ec2.tbl_real_fund_txn_type_mst colunas: {df_mst['column_name'].tolist()}")

    return True


# ==============================================================================
# STEP 1: SQL Principal — Cohort por dia de registro
# ==============================================================================
# NOTA SOBRE PERFORMANCE:
#   fund_ec2.tbl_real_fund_txn NAO tem coluna de particao (dt).
#   Para evitar full scan, filtramos c_start_time >= DATE_START.
#   Mesmo assim, como precisamos da PRIMEIRA aposta de cada jogador
#   (que pode ser anterior ao periodo), a CTE first_bets nao tem filtro
#   de data -- ela pega o MIN global. Isso pode ser pesado.
#
#   OTIMIZACAO: Como estamos analisando jogadores que se REGISTRARAM nos
#   ultimos 30 dias, a first bet deles so pode ter acontecido apos o
#   registro. Entao filtramos fund_ec2 por c_start_time >= DATE_START
#   (mesma janela dos registros). Isso reduz drasticamente o scan.

MAIN_SQL = f"""
-- ================================================================
-- Conversao Registro -> Primeira Aposta Gaming (Cohort)
-- Periodo de registro: {DATE_START} a {DATE_END} (BRT)
-- Perspectiva: para cada dia de registro, % que ja fez first bet
-- ================================================================

WITH
-- ----------------------------------------------------------------
-- CTE 1: Registros no periodo (BRT)
-- Fonte: ecr_ec2.tbl_ecr + ecr_ec2.tbl_ecr_flags
-- Filtro: c_test_user = false, signup dentro do periodo
-- ----------------------------------------------------------------
registros AS (
    SELECT
        e.c_ecr_id,
        -- Converter UTC -> BRT para agrupar por dia local
        e.c_signup_time AT TIME ZONE 'UTC' AT TIME ZONE 'America/Sao_Paulo'
            AS signup_brt,
        CAST(
            e.c_signup_time AT TIME ZONE 'UTC' AT TIME ZONE 'America/Sao_Paulo'
            AS DATE
        ) AS reg_date_brt
    FROM ecr_ec2.tbl_ecr e
    -- Excluir test users (obrigatorio)
    JOIN ecr_ec2.tbl_ecr_flags f ON e.c_ecr_id = f.c_ecr_id
    WHERE f.c_test_user = false
      -- Filtrar por periodo de registro (em UTC, com margem de 1 dia
      -- para cobrir a conversao UTC->BRT no limite)
      AND e.c_signup_time >= TIMESTAMP '{DATE_START} 03:00:00'
      AND e.c_signup_time < TIMESTAMP '{DATE_END}' + INTERVAL '1' DAY + INTERVAL '3' HOUR
),

-- ----------------------------------------------------------------
-- CTE 2: Primeira aposta gaming de cada jogador registrado no periodo
-- Fonte: fund_ec2.tbl_real_fund_txn
-- Filtro: gaming (via type_mst), debito (op_type=DB), status=SUCCESS
-- OTIMIZACAO: so olhar transacoes apos a data de inicio do periodo
-- (jogadores novos nao podem ter apostado antes de existir)
-- ----------------------------------------------------------------
first_bets AS (
    SELECT
        t.c_ecr_id,
        -- MIN pega a primeira aposta de cada jogador
        MIN(
            t.c_start_time AT TIME ZONE 'UTC' AT TIME ZONE 'America/Sao_Paulo'
        ) AS first_bet_brt
    FROM fund_ec2.tbl_real_fund_txn t
    -- Classificar como gaming via tabela mestre
    JOIN fund_ec2.tbl_real_fund_txn_type_mst m
        ON t.c_txn_type = m.c_txn_type
    -- Excluir test users
    JOIN ecr_ec2.tbl_ecr_flags f
        ON t.c_ecr_id = f.c_ecr_id
    WHERE m.c_is_gaming_txn = 'Y'       -- so transacoes gaming
      AND t.c_op_type = 'DB'            -- debito = aposta
      AND t.c_txn_status = 'SUCCESS'    -- so confirmadas
      AND f.c_test_user = false
      -- OTIMIZACAO: limitar scan — jogadores novos nao apostaram antes
      AND t.c_start_time >= TIMESTAMP '{DATE_START} 03:00:00'
    GROUP BY t.c_ecr_id
)

-- ----------------------------------------------------------------
-- Query final: join cohort por dia de registro
-- LEFT JOIN para incluir registros SEM first bet (nao convertidos)
-- ----------------------------------------------------------------
SELECT
    r.reg_date_brt,
    -- Dia da semana (1=Segunda, 7=Domingo no Presto)
    day_of_week(r.reg_date_brt) AS dow_number,
    CASE day_of_week(r.reg_date_brt)
        WHEN 1 THEN 'Segunda'
        WHEN 2 THEN 'Terca'
        WHEN 3 THEN 'Quarta'
        WHEN 4 THEN 'Quinta'
        WHEN 5 THEN 'Sexta'
        WHEN 6 THEN 'Sabado'
        WHEN 7 THEN 'Domingo'
    END AS dia_semana,
    -- Quantidade de registros no dia
    COUNT(DISTINCT r.c_ecr_id) AS qty_registrations,
    -- Quantidade que ja fez first bet
    COUNT(DISTINCT fb.c_ecr_id) AS qty_first_bets,
    -- Taxa de conversao (%)
    ROUND(
        CAST(COUNT(DISTINCT fb.c_ecr_id) AS DOUBLE)
        / NULLIF(CAST(COUNT(DISTINCT r.c_ecr_id) AS DOUBLE), 0)
        * 100,
        2
    ) AS conversion_rate_pct,
    -- Tempo medio entre registro e first bet (horas)
    ROUND(
        AVG(
            CASE
                WHEN fb.first_bet_brt IS NOT NULL
                THEN date_diff('minute', r.signup_brt, fb.first_bet_brt) / 60.0
            END
        ),
        2
    ) AS avg_hours_to_first_bet,
    -- Mediana aprox: quantos converteram em D+0 (mesmo dia)
    COUNT_IF(
        fb.first_bet_brt IS NOT NULL
        AND CAST(fb.first_bet_brt AS DATE) = r.reg_date_brt
    ) AS qty_same_day_bets,
    -- Quantos converteram em D+1
    COUNT_IF(
        fb.first_bet_brt IS NOT NULL
        AND CAST(fb.first_bet_brt AS DATE) = date_add('day', 1, r.reg_date_brt)
    ) AS qty_d1_bets,
    -- Quantos converteram em D+2 a D+7
    COUNT_IF(
        fb.first_bet_brt IS NOT NULL
        AND CAST(fb.first_bet_brt AS DATE) > date_add('day', 1, r.reg_date_brt)
        AND CAST(fb.first_bet_brt AS DATE) <= date_add('day', 7, r.reg_date_brt)
    ) AS qty_d2_d7_bets,
    -- Quantos converteram depois de D+7
    COUNT_IF(
        fb.first_bet_brt IS NOT NULL
        AND CAST(fb.first_bet_brt AS DATE) > date_add('day', 7, r.reg_date_brt)
    ) AS qty_after_d7_bets
FROM registros r
LEFT JOIN first_bets fb ON r.c_ecr_id = fb.c_ecr_id
WHERE r.reg_date_brt BETWEEN DATE '{DATE_START}' AND DATE '{DATE_END}'
GROUP BY r.reg_date_brt, day_of_week(r.reg_date_brt)
ORDER BY r.reg_date_brt
"""


# ==============================================================================
# STEP 2: Executar query e processar
# ==============================================================================
def run_query():
    """Executa a query principal no Athena e retorna DataFrame."""
    log.info(f"=== STEP 1: Executando query principal ===")
    log.info(f"Periodo: {DATE_START} a {DATE_END}")
    log.info("AVISO: Query pode demorar ~2-5min (fund_ec2 sem particao)")
    log.info("SQL sendo executado:")
    log.info(MAIN_SQL)

    df = query_athena(MAIN_SQL, database="fund_ec2")
    log.info(f"Resultado: {len(df)} dias retornados")

    if len(df) == 0:
        log.error("ATENCAO: Query retornou 0 linhas! Verificar filtros.")
        raise ValueError("Query retornou 0 linhas")

    # Garantir tipos corretos
    df['reg_date_brt'] = pd.to_datetime(df['reg_date_brt']).dt.date
    for col in ['qty_registrations', 'qty_first_bets', 'qty_same_day_bets',
                'qty_d1_bets', 'qty_d2_d7_bets', 'qty_after_d7_bets']:
        df[col] = df[col].astype(int)
    df['conversion_rate_pct'] = df['conversion_rate_pct'].astype(float)
    df['avg_hours_to_first_bet'] = df['avg_hours_to_first_bet'].astype(float)

    return df


# ==============================================================================
# STEP 3: Calcular metricas derivadas em Python
# ==============================================================================
def enrich_dataframe(df):
    """Adiciona media movel 7 dias e metricas derivadas."""
    log.info("=== STEP 2: Enriquecendo dados ===")

    # Media movel 7 dias da taxa de conversao
    df['conversion_ma7'] = (
        df['conversion_rate_pct']
        .rolling(window=7, min_periods=1)
        .mean()
        .round(2)
    )

    # Percentual de conversao same-day sobre total de first bets
    df['pct_same_day'] = (
        df.apply(
            lambda r: round(r['qty_same_day_bets'] / r['qty_first_bets'] * 100, 1)
            if r['qty_first_bets'] > 0 else 0.0,
            axis=1
        )
    )

    # "Dias de maturacao" — cohorts recentes tem menos tempo para converter
    # Adicionar flag de maturacao
    df['days_matured'] = df['reg_date_brt'].apply(
        lambda d: (TODAY.date() - d).days
    )

    return df


# ==============================================================================
# STEP 4: Analise e insights (print no console)
# ==============================================================================
def print_analysis(df):
    """Imprime analise e insights no console."""
    print("\n" + "=" * 80)
    print("CONVERSAO REGISTRO -> PRIMEIRA APOSTA GAMING (ULTIMOS 30 DIAS)")
    print("=" * 80)
    print(f"\nPeriodo: {DATE_START} a {DATE_END}")
    print(f"Data da analise: {TODAY.strftime('%Y-%m-%d')}")
    print(f"Perspectiva: Cohort por dia de registro\n")

    # Tabela diaria
    display_cols = [
        'reg_date_brt', 'dia_semana', 'qty_registrations', 'qty_first_bets',
        'conversion_rate_pct', 'conversion_ma7', 'avg_hours_to_first_bet',
        'qty_same_day_bets', 'days_matured'
    ]
    print(df[display_cols].to_string(index=False))

    # Metricas globais
    total_regs = df['qty_registrations'].sum()
    total_bets = df['qty_first_bets'].sum()
    taxa_global = round(total_bets / total_regs * 100, 2) if total_regs > 0 else 0
    avg_hours = df['avg_hours_to_first_bet'].mean()
    same_day_total = df['qty_same_day_bets'].sum()

    print(f"\n{'='*80}")
    print("METRICAS GLOBAIS")
    print(f"{'='*80}")
    print(f"  Total de registros:          {total_regs:,}")
    print(f"  Total de first bets:         {total_bets:,}")
    print(f"  Taxa de conversao global:    {taxa_global}%")
    print(f"  Tempo medio -> first bet:    {avg_hours:.1f} horas")
    print(f"  Conversoes same-day (D+0):   {same_day_total:,} ({round(same_day_total/total_bets*100,1) if total_bets else 0}% das first bets)")
    print(f"  Media diaria de registros:   {total_regs/len(df):.0f}")
    print(f"  Media diaria de first bets:  {total_bets/len(df):.0f}")

    # Analise por dia da semana
    print(f"\n{'='*80}")
    print("ANALISE POR DIA DA SEMANA")
    print(f"{'='*80}")
    dow_agg = df.groupby('dia_semana').agg(
        dias=('reg_date_brt', 'count'),
        avg_regs=('qty_registrations', 'mean'),
        avg_bets=('qty_first_bets', 'mean'),
        avg_conv=('conversion_rate_pct', 'mean')
    ).round(1)

    # Reordenar dias da semana
    order = ['Segunda', 'Terca', 'Quarta', 'Quinta', 'Sexta', 'Sabado', 'Domingo']
    dow_agg = dow_agg.reindex([d for d in order if d in dow_agg.index])
    print(dow_agg.to_string())

    # Tendencia: comparar 1a metade vs 2a metade do periodo
    mid = len(df) // 2
    first_half = df.iloc[:mid]
    second_half = df.iloc[mid:]
    conv_1h = round(
        first_half['qty_first_bets'].sum() / first_half['qty_registrations'].sum() * 100, 2
    ) if first_half['qty_registrations'].sum() > 0 else 0
    conv_2h = round(
        second_half['qty_first_bets'].sum() / second_half['qty_registrations'].sum() * 100, 2
    ) if second_half['qty_registrations'].sum() > 0 else 0

    print(f"\n{'='*80}")
    print("TENDENCIA (1a metade vs 2a metade do periodo)")
    print(f"{'='*80}")
    print(f"  1a metade ({first_half['reg_date_brt'].iloc[0]} a {first_half['reg_date_brt'].iloc[-1]}):")
    print(f"    Registros: {first_half['qty_registrations'].sum():,} | First Bets: {first_half['qty_first_bets'].sum():,} | Conversao: {conv_1h}%")
    print(f"  2a metade ({second_half['reg_date_brt'].iloc[0]} a {second_half['reg_date_brt'].iloc[-1]}):")
    print(f"    Registros: {second_half['qty_registrations'].sum():,} | First Bets: {second_half['qty_first_bets'].sum():,} | Conversao: {conv_2h}%")
    delta = conv_2h - conv_1h
    direction = "MELHORIA" if delta > 0 else "QUEDA" if delta < 0 else "ESTAVEL"
    print(f"  Variacao: {delta:+.2f}pp ({direction})")

    # Alerta sobre cohorts recentes
    recent = df[df['days_matured'] <= 3]
    if len(recent) > 0:
        print(f"\n  AVISO: Os ultimos {len(recent)} dias tem menos de 3 dias de")
        print(f"  maturacao. A taxa de conversao deles PODE subir ainda.")
        print(f"  Para comparacao justa, considere apenas cohorts com 7+ dias.")

    # Acoes sugeridas
    print(f"\n{'='*80}")
    print("ACOES SUGERIDAS")
    print(f"{'='*80}")
    print("""
    1. IDENTIFICAR DATA DA MUDANCA NO CADASTRO:
       - Pedir ao time de produto a data exata da mudanca
       - Marcar no grafico e comparar antes/depois

    2. SE CONVERSAO MELHOROU:
       - Documentar o impacto para apresentar ao CTO/CGO
       - Monitorar se a melhoria se mantem nas proximas semanas

    3. SE CONVERSAO NAO MUDOU OU PIOROU:
       - Investigar se o novo fluxo criou friccao inesperada
       - Verificar se houve mudanca no mix de canais de aquisicao
       - Cruzar com dados de CRM (automacoes de boas-vindas)

    4. OTIMIZACAO DO FUNIL:
       - Jogadores que registram e NAO apostam em D+0: enviar push/SMS
       - D+1 sem aposta: email com tutorial + bonus de first bet
       - D+7 sem aposta: considerar como lost, focar em reativacao

    5. PROXIMOS PASSOS:
       - Cruzar com canais de aquisicao (qual canal converte melhor?)
       - Segmentar por device (mobile vs desktop)
       - Analisar se first bet leva a first deposit ou vice-versa
""")

    return {
        'total_regs': total_regs,
        'total_bets': total_bets,
        'taxa_global': taxa_global,
        'avg_hours': avg_hours,
        'conv_1h': conv_1h,
        'conv_2h': conv_2h,
        'delta': delta,
    }


# ==============================================================================
# STEP 5: Salvar CSV
# ==============================================================================
def save_csv(df):
    """Salva DataFrame em CSV."""
    log.info(f"=== STEP 3: Salvando CSV ===")
    df.to_csv(CSV_PATH, index=False, encoding='utf-8-sig')
    log.info(f"CSV salvo: {CSV_PATH}")
    return CSV_PATH


# ==============================================================================
# STEP 6: Gerar grafico (matplotlib)
# ==============================================================================
def generate_chart(df, metrics):
    """Gera grafico temporal com barras + linha de conversao."""
    log.info("=== STEP 4: Gerando grafico ===")

    fig, ax1 = plt.subplots(figsize=(16, 8))

    dates = pd.to_datetime(df['reg_date_brt'])
    x = range(len(dates))
    width = 0.35

    # Barras: registros (azul) e first bets (verde)
    bars1 = ax1.bar(
        [i - width/2 for i in x],
        df['qty_registrations'],
        width,
        label='Registros',
        color='#4A90D9',
        alpha=0.8,
        edgecolor='white',
        linewidth=0.5
    )
    bars2 = ax1.bar(
        [i + width/2 for i in x],
        df['qty_first_bets'],
        width,
        label='First Bets',
        color='#2ECC71',
        alpha=0.8,
        edgecolor='white',
        linewidth=0.5
    )

    ax1.set_xlabel('Data de Registro (BRT)', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Quantidade', fontsize=12, fontweight='bold', color='#333333')
    ax1.tick_params(axis='y', labelcolor='#333333')
    ax1.yaxis.set_major_locator(MaxNLocator(integer=True))

    # Eixo X com datas
    ax1.set_xticks(list(x))
    ax1.set_xticklabels(
        [d.strftime('%d/%m') for d in dates],
        rotation=45,
        ha='right',
        fontsize=8
    )

    # Eixo Y direito: taxa de conversao
    ax2 = ax1.twinx()
    ax2.plot(
        list(x),
        df['conversion_rate_pct'],
        color='#E74C3C',
        linewidth=1.5,
        marker='o',
        markersize=4,
        label='Taxa Conversao (%)',
        alpha=0.8
    )
    # Media movel 7 dias
    ax2.plot(
        list(x),
        df['conversion_ma7'],
        color='#E74C3C',
        linewidth=2.5,
        linestyle='--',
        label='MM7 Conversao (%)',
        alpha=0.6
    )

    ax2.set_ylabel('Taxa de Conversao (%)', fontsize=12, fontweight='bold', color='#E74C3C')
    ax2.tick_params(axis='y', labelcolor='#E74C3C')

    # Titulo
    fig.suptitle(
        'Conversao Registro -> Primeira Aposta Gaming',
        fontsize=16,
        fontweight='bold',
        y=0.98
    )
    ax1.set_title(
        f'Periodo: {DATE_START} a {DATE_END} | '
        f'Taxa Global: {metrics["taxa_global"]}% | '
        f'Tendencia: {metrics["delta"]:+.2f}pp',
        fontsize=10,
        color='#666666',
        pad=10
    )

    # Legenda combinada
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(
        lines1 + lines2,
        labels1 + labels2,
        loc='upper left',
        framealpha=0.9,
        fontsize=9
    )

    # Linhas de referencia
    avg_conv = df['conversion_rate_pct'].mean()
    ax2.axhline(
        y=avg_conv,
        color='#E74C3C',
        linestyle=':',
        alpha=0.3,
        linewidth=1
    )
    ax2.annotate(
        f'Media: {avg_conv:.1f}%',
        xy=(len(x)-1, avg_conv),
        fontsize=8,
        color='#E74C3C',
        alpha=0.5
    )

    # Linha vertical marcando mudanca no fluxo de cadastro (18/03/2026)
    change_date = pd.Timestamp('2026-03-18')
    if change_date >= dates.iloc[0] and change_date <= dates.iloc[-1]:
        change_idx = None
        for i, d in enumerate(dates):
            if d.date() == change_date.date():
                change_idx = i
                break
        if change_idx is not None:
            ax1.axvline(x=change_idx, color='#9B59B6', linewidth=2, linestyle='--', alpha=0.7)
            ax1.annotate(
                'Mudanca cadastro\n(auto endereco)',
                xy=(change_idx, df['qty_registrations'].max()),
                xytext=(change_idx - 3, df['qty_registrations'].max() * 1.05),
                fontsize=8,
                fontweight='bold',
                color='#9B59B6',
                arrowprops=dict(arrowstyle='->', color='#9B59B6'),
                bbox=dict(boxstyle='round,pad=0.3', facecolor='#F5EEF8', edgecolor='#9B59B6')
            )

    # Aviso sobre maturacao dos cohorts recentes
    ax1.annotate(
        'Cohorts recentes podem\nainda nao ter maturado',
        xy=(len(x)-2, df['qty_registrations'].iloc[-2]),
        xytext=(len(x)-8, df['qty_registrations'].max() * 0.9),
        fontsize=7,
        color='#999999',
        arrowprops=dict(arrowstyle='->', color='#cccccc'),
        bbox=dict(boxstyle='round,pad=0.3', facecolor='#f0f0f0', edgecolor='#cccccc')
    )

    plt.tight_layout()
    fig.savefig(PNG_PATH, dpi=150, bbox_inches='tight')
    plt.close(fig)
    log.info(f"Grafico salvo: {PNG_PATH}")
    return PNG_PATH


# ==============================================================================
# STEP 7: Gerar Excel com dados + legenda
# ==============================================================================
def generate_excel(df, metrics):
    """Gera Excel com aba de dados, aba de legenda e grafico embedded."""
    log.info("=== STEP 5: Gerando Excel ===")

    wb = openpyxl.Workbook()

    # ---- ABA 1: Dados ----
    ws_data = wb.active
    ws_data.title = "Dados"

    # Header com contexto
    ws_data.append(["Conversao Registro -> Primeira Aposta Gaming"])
    ws_data.append([f"Periodo: {DATE_START} a {DATE_END}"])
    ws_data.append([f"Gerado em: {TODAY.strftime('%Y-%m-%d')}"])
    ws_data.append([])  # linha vazia

    # Dados do DataFrame
    headers = [
        'Data Registro', 'DoW', 'Dia Semana', 'Registros', 'First Bets',
        'Conversao %', 'MM7 Conv %', 'Horas p/ First Bet',
        'Same Day (D+0)', 'D+1', 'D+2 a D+7', 'Depois D+7',
        '% Same Day', 'Dias Maturacao'
    ]
    ws_data.append(headers)

    for _, row in df.iterrows():
        ws_data.append([
            str(row['reg_date_brt']),
            row['dow_number'],
            row['dia_semana'],
            row['qty_registrations'],
            row['qty_first_bets'],
            row['conversion_rate_pct'],
            row['conversion_ma7'],
            row['avg_hours_to_first_bet'],
            row['qty_same_day_bets'],
            row['qty_d1_bets'],
            row['qty_d2_d7_bets'],
            row['qty_after_d7_bets'],
            row['pct_same_day'],
            row['days_matured']
        ])

    # Resumo no final
    row_start = len(df) + 7
    ws_data.append([])
    ws_data.append(["RESUMO"])
    ws_data.append(["Total Registros", metrics['total_regs']])
    ws_data.append(["Total First Bets", metrics['total_bets']])
    ws_data.append(["Taxa Conversao Global %", metrics['taxa_global']])
    ws_data.append(["Tempo Medio -> First Bet (h)", round(metrics['avg_hours'], 1)])
    ws_data.append(["Conv 1a Metade %", metrics['conv_1h']])
    ws_data.append(["Conv 2a Metade %", metrics['conv_2h']])
    ws_data.append(["Variacao (pp)", round(metrics['delta'], 2)])

    # Formatar largura de colunas
    for i, h in enumerate(headers, 1):
        ws_data.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = max(len(h) + 2, 14)

    # ---- Grafico embutido no Excel ----
    try:
        from openpyxl.drawing.image import Image as XlImage
        img = XlImage(PNG_PATH)
        img.width = 900
        img.height = 450
        ws_data.add_image(img, f"A{row_start + 10}")
        log.info("Grafico PNG embutido no Excel")
    except Exception as e:
        log.warning(f"Nao foi possivel embutir grafico no Excel: {e}")

    # ---- ABA 2: Legenda ----
    ws_leg = wb.create_sheet("Legenda")

    legenda_rows = [
        ["LEGENDA - Conversao Registro -> Primeira Aposta Gaming"],
        [],
        ["INFORMACOES GERAIS"],
        ["Data da analise", TODAY.strftime('%Y-%m-%d')],
        ["Periodo analisado", f"{DATE_START} a {DATE_END}"],
        ["Fonte dos registros", "ecr_ec2.tbl_ecr (Athena, Iceberg Data Lake)"],
        ["Fonte das apostas", "fund_ec2.tbl_real_fund_txn (Athena, Iceberg Data Lake)"],
        ["Fuso horario", "BRT (America/Sao_Paulo) - convertido de UTC"],
        ["Filtros aplicados", "c_test_user = false (excluindo usuarios de teste)"],
        ["Perspectiva", "Cohort por dia de REGISTRO (nao por dia de aposta)"],
        [],
        ["COLUNAS"],
        ["Data Registro", "Dia em que o jogador se registrou (BRT)"],
        ["DoW", "Numero do dia da semana (1=Segunda, 7=Domingo)"],
        ["Dia Semana", "Nome do dia da semana"],
        ["Registros", "Qtd de jogadores que se registraram nesse dia"],
        ["First Bets", "Qtd de jogadores desse cohort que ja fizeram sua 1a aposta gaming"],
        ["Conversao %", "(First Bets / Registros) * 100"],
        ["MM7 Conv %", "Media movel 7 dias da taxa de conversao (suaviza oscilacoes)"],
        ["Horas p/ First Bet", "Tempo medio (em horas) entre registro e 1a aposta"],
        ["Same Day (D+0)", "Quantos jogadores apostaram no MESMO DIA do registro"],
        ["D+1", "Quantos apostaram no dia seguinte ao registro"],
        ["D+2 a D+7", "Quantos apostaram entre 2 e 7 dias apos registro"],
        ["Depois D+7", "Quantos apostaram mais de 7 dias apos registro"],
        ["% Same Day", "(Same Day / First Bets) * 100 - proporcao de conversao imediata"],
        ["Dias Maturacao", "Dias desde o registro ate hoje. Cohorts recentes (<7 dias) podem ter taxa artificialmente menor"],
        [],
        ["GLOSSARIO"],
        ["First Bet", "Primeira aposta GAMING (casino + sportsbook). NAO e primeiro deposito (FTD)."],
        ["Cohort", "Grupo de jogadores agrupados pelo dia de registro."],
        ["Conversao", "% de jogadores registrados que fizeram first bet em qualquer momento posterior."],
        ["MM7", "Media movel de 7 dias. Suaviza variacao diaria para mostrar tendencia."],
        ["Maturacao", "Tempo que o cohort teve para converter. Cohorts muito recentes parecem piores porque ainda nao tiveram tempo."],
        ["D+0, D+1, etc.", "Dias apos o registro. D+0 = mesmo dia, D+1 = dia seguinte."],
        [],
        ["COMO INTERPRETAR"],
        ["Conversao subindo", "Mudanca no cadastro pode estar melhorando. Validar com mais dados."],
        ["Conversao estavel", "Mudanca no cadastro nao teve impacto significativo."],
        ["Conversao caindo", "Possivel friccao no novo fluxo. Investigar com time de produto."],
        ["Cohorts recentes com conversao menor", "NORMAL - eles ainda nao tiveram tempo de converter. Comparar apenas cohorts com 7+ dias."],
        ["% Same Day alto", "Jogadores convertem rapido. Bom sinal de UX."],
        ["Horas p/ First Bet diminuindo", "Friccao no funil esta diminuindo (positivo)."],
        [],
        ["ACAO SUGERIDA"],
        ["Se conversao melhorou", "Apresentar resultado ao CTO/CGO como case de otimizacao."],
        ["Se conversao caiu", "Levantar hipoteses com produto e CRM. Verificar se ha friccao nova."],
        ["Jogadores D+0 sem bet", "CRM: push/SMS imediato com tutorial e incentivo."],
        ["Jogadores D+1 sem bet", "CRM: email com bonus de first bet."],
        ["Jogadores D+7 sem bet", "Classificar como 'lost'. Focar em reativacao com desconto."],
    ]

    for row in legenda_rows:
        ws_leg.append(row)

    ws_leg.column_dimensions['A'].width = 35
    ws_leg.column_dimensions['B'].width = 80

    # Salvar
    wb.save(XLSX_PATH)
    log.info(f"Excel salvo: {XLSX_PATH}")
    return XLSX_PATH


# ==============================================================================
# MAIN
# ==============================================================================
def main():
    log.info("=" * 80)
    log.info("INICIO: Conversao Registro -> Primeira Aposta Gaming (30 dias)")
    log.info("=" * 80)

    # Step 0: Validar schemas
    try:
        discover_schemas()
    except Exception as e:
        log.warning(f"Validacao de schema falhou: {e}")
        log.info("Continuando mesmo assim -- pode ser problema de permissao em information_schema")

    # Step 1: Executar query
    df = run_query()

    # Step 2: Enriquecer dados
    df = enrich_dataframe(df)

    # Step 3: Analise e insights
    metrics = print_analysis(df)

    # Step 4: Salvar CSV
    save_csv(df)

    # Step 5: Gerar grafico
    generate_chart(df, metrics)

    # Step 6: Gerar Excel
    generate_excel(df, metrics)

    # Resumo final
    log.info("=" * 80)
    log.info("CONCLUIDO!")
    log.info(f"  CSV:     {CSV_PATH}")
    log.info(f"  Grafico: {PNG_PATH}")
    log.info(f"  Excel:   {XLSX_PATH}")
    log.info("=" * 80)


if __name__ == "__main__":
    main()

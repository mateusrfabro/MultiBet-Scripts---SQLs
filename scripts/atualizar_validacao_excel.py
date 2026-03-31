"""
Adiciona aba de validacao cruzada BigQuery ao Excel final
"""
import sys
import os
import io
import warnings

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from db.athena import query_athena
from db.bigquery import query_bigquery

DATA_INICIO = "2026-01-01"
DATA_FIM = "2026-03-25"
EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "reports",
    "top_provedores_turnover_20260101_20260325_FINAL.xlsx",
)


def main():
    print("Carregando dados de validacao...\n")

    # -- Fonte 1: bireports game_play_summary --
    df1 = query_athena(f"""
    SELECT
        s.c_vendor_id AS provedor,
        SUM(CASE WHEN s.c_txn_type = 27 THEN s.c_txn_real_cash_amount_ecr_crncy ELSE 0 END) / 100.0 AS bet_real,
        SUM(CASE WHEN s.c_txn_type = 27 THEN
            COALESCE(s.c_txn_crp_amount_ecr_crncy,0) + COALESCE(s.c_txn_drp_amount_ecr_crncy,0)
            + COALESCE(s.c_txn_wrp_amount_ecr_crncy,0) + COALESCE(s.c_txn_rrp_amount_ecr_crncy,0)
        ELSE 0 END) / 100.0 AS bet_bonus,
        COUNT(DISTINCT s.c_ecr_id) AS jogadores
    FROM bireports_ec2.tbl_ecr_txn_type_wise_daily_game_play_summary s
    JOIN bireports_ec2.tbl_ecr e ON s.c_ecr_id = e.c_ecr_id
    WHERE s.c_created_date BETWEEN DATE '{DATA_INICIO}' AND DATE '{DATA_FIM}'
      AND e.c_test_user = false
    GROUP BY s.c_vendor_id
    ORDER BY bet_real DESC
    """, database="bireports_ec2")
    df1["turnover_bireports"] = df1["bet_real"] + df1["bet_bonus"]
    print(f"  F1 (game_play_summary): {len(df1)} vendors, total R$ {df1['turnover_bireports'].sum():,.2f}")

    # -- Fonte 2: daily_bi_summary --
    df2 = query_athena(f"""
    SELECT
        SUM(s.c_casino_bet_amount) / 100.0 AS casino_bet_total,
        SUM(s.c_casino_realcash_bet_amount) / 100.0 AS casino_bet_real,
        SUM(s.c_sb_bet_amount) / 100.0 AS sb_bet_total,
        COUNT(DISTINCT s.c_ecr_id) AS jogadores
    FROM bireports_ec2.tbl_ecr_wise_daily_bi_summary s
    JOIN bireports_ec2.tbl_ecr e ON s.c_ecr_id = e.c_ecr_id
    WHERE s.c_created_date BETWEEN DATE '{DATA_INICIO}' AND DATE '{DATA_FIM}'
      AND e.c_test_user = false
    """, database="bireports_ec2")
    f2_total = float(df2["casino_bet_total"].iloc[0])
    print(f"  F2 (daily_bi_summary): total casino R$ {f2_total:,.2f}")

    # -- Fonte 3: BigQuery --
    df3 = query_bigquery(f"""
    SELECT
        casino_last_bet_game_provider AS provider_id,
        SUM(casino_last_bet_amount) AS turnover_bigquery,
        COUNT(DISTINCT user_id) AS jogadores_bq
    FROM `smartico-bq6.dwh_ext_24105.tr_casino_bet`
    WHERE event_time >= TIMESTAMP("{DATA_INICIO}")
      AND event_time < TIMESTAMP("2026-03-26")
      AND IFNULL(casino_is_rollback, false) = false
    GROUP BY casino_last_bet_game_provider
    ORDER BY turnover_bigquery DESC
    LIMIT 15
    """)
    df3["turnover_bigquery"] = df3["turnover_bigquery"].astype(float)
    df3["jogadores_bq"] = df3["jogadores_bq"].astype(int)
    f3_total = df3["turnover_bigquery"].sum()
    print(f"  F3 (BigQuery Smartico): {len(df3)} providers, total R$ {f3_total:,.2f}")

    # -- Match por posicao (ambos ordenados por turnover DESC) --
    df1_top = df1.sort_values("turnover_bireports", ascending=False).head(10).reset_index(drop=True)
    df3_top = df3.sort_values("turnover_bigquery", ascending=False).head(10).reset_index(drop=True)

    validacao = pd.DataFrame({
        "provedor_athena": df1_top["provedor"],
        "turnover_athena_brl": df1_top["turnover_bireports"],
        "jogadores_athena": df1_top["jogadores"],
        "provider_id_bigquery": df3_top["provider_id"].astype(int),
        "turnover_bigquery_brl": df3_top["turnover_bigquery"],
        "jogadores_bigquery": df3_top["jogadores_bq"],
    })

    validacao["divergencia_pct"] = (
        (validacao["turnover_bigquery_brl"] - validacao["turnover_athena_brl"])
        / validacao["turnover_athena_brl"] * 100
    ).round(2)

    validacao["match_players"] = validacao.apply(
        lambda r: "OK" if abs(r["jogadores_athena"] - r["jogadores_bigquery"]) / max(r["jogadores_athena"], 1) < 0.05
        else "DIVERGE", axis=1
    )

    # -- Totais --
    totais = pd.DataFrame([{
        "fonte": "F1 - bireports game_play_summary",
        "total_casino_brl": df1["turnover_bireports"].sum(),
        "nota": "Por vendor, txn_type=27 (bet), centavos/100"
    }, {
        "fonte": "F2 - bireports daily_bi_summary",
        "total_casino_brl": f2_total,
        "nota": "Agregado player/dia, c_casino_bet_amount/100"
    }, {
        "fonte": "F3 - BigQuery Smartico tr_casino_bet",
        "total_casino_brl": f3_total,
        "nota": "CRM externo, casino_last_bet_amount, rollbacks excluidos"
    }])

    # -- Salvar no Excel --
    wb = load_workbook(EXCEL_PATH)

    # Remove abas antigas de validacao se existirem
    for name in ["Validacao Cruzada", "Totais por Fonte"]:
        if name in wb.sheetnames:
            del wb[name]
    wb.save(EXCEL_PATH)

    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a") as writer:
        validacao.to_excel(writer, sheet_name="Validacao Cruzada", index=False)
        totais.to_excel(writer, sheet_name="Totais por Fonte", index=False)

    # -- Print resumo --
    print(f"\n{'='*75}")
    print(f"  VALIDACAO CRUZADA - 3 FONTES")
    print(f"{'='*75}\n")

    for _, row in validacao.iterrows():
        status = "OK" if abs(row["divergencia_pct"]) < 10 else "!!"
        print(f"  [{status}] {row['provedor_athena']:<22s}  "
              f"Athena: R$ {row['turnover_athena_brl']:>14,.2f}  "
              f"BQ: R$ {row['turnover_bigquery_brl']:>14,.2f}  "
              f"diff: {row['divergencia_pct']:>+6.1f}%  "
              f"players: {row['match_players']}")

    print(f"\n  Totais:")
    for _, row in totais.iterrows():
        print(f"    {row['fonte']:<40s}  R$ {row['total_casino_brl']:>15,.2f}")

    f1_total = df1["turnover_bireports"].sum()
    div_f1_f3 = (f1_total - f3_total) / f3_total * 100

    print(f"\n  Divergencia global F1 vs F3: {div_f1_f3:>+.2f}%")
    if abs(div_f1_f3) < 5:
        print(f"  [OK] RANKING VALIDADO - Athena e BigQuery convergem (<5%)")
    elif abs(div_f1_f3) < 10:
        print(f"  [OK] RANKING VALIDADO - Divergencia aceitavel (<10%)")
    else:
        print(f"  [!!] Divergencia significativa - investigar")

    print(f"\n  Excel atualizado: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
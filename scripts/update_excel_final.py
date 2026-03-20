"""
Atualiza Excel com KPIs que eram PENDENTES e agora estao resolvidos.
Tambem adiciona fact_redeposits (nova tabela, fonte verde).
"""
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
ws = wb['Mapa de KPIs']

green_font = Font(color="00B050", bold=False)

def set_cells(row, h=None, i=None, j=None, k=None, green=False):
    for col, val in [('H', h), ('I', i), ('J', j), ('K', k)]:
        if val:
            cell = ws[f'{col}{row}']
            cell.value = val
            if green:
                cell.font = green_font

# ============================================================
# ATUALIZAR PENDENTES -> IMPLEMENTADOS
# ============================================================

# L41: FTD por metodo pagamento
set_cells(41, 'IMPLEMENTADA: 99.99% dos FTDs via pay2free_direct_multibet (PIX). Sem valor analitico - apenas 1 PSP operando', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (c_processor_name)')

# L43: % FTD com bonus
set_cells(43, 'IMPLEMENTADA: 0.48% dos FTDs tiveram bonus ativo (748 de 154.982). Bonus quase inexistente no momento do FTD', 'Athena', 'bonus_ec2', 'tbl_ecr_bonus_details (c_ecr_id, c_bonus_status) LEFT JOIN cashier_ec2')

# L58: LTV/CAC ratio
set_cells(58, 'IMPLEMENTADA via vw_ltv_cac_ratio: AVG(ggr_d30) / CAC. View com month_of_ftd x source', 'Calculado', 'multibet', 'vw_ltv_cac_ratio (avg_ltv_d30, cac, ltv_cac_ratio)')

# L73: Duracao media sessao
set_cells(73, 'IMPLEMENTADA: AVG(date_diff(minute, c_start_time, c_updated_time)) filtro 1-480min. Avg ~100min/dia', 'Athena', 'fund_ec2', 'tbl_real_fund_session (c_start_time, c_updated_time)')

# L75: Taxa 2o deposito D7 (update com dados da fact_redeposits)
set_cells(75, 'IMPLEMENTADA via fact_redeposits: is_redepositor_d7. Global: 39.3% fazem 2o deposito em 7 dias. 46.9% total redepositors', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (ROW_NUMBER rn=2, date_diff <= 7)')

# L76: Frequencia de redeposit
set_cells(76, 'IMPLEMENTADA via fact_redeposits: deposits_per_month = total_deposits / meses desde FTD', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (COUNT por player / meses ativo)')

# L77: Ticket medio redeposit
set_cells(77, 'IMPLEMENTADA via fact_redeposits: avg_redeposit_amount = AVG(amount) WHERE rn > 1. Avg: R$ 190.24', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (c_confirmed_amount WHERE deposit_rank > 1)')

# L78: Intervalo medio entre depositos
set_cells(78, 'IMPLEMENTADA via fact_redeposits: avg_days_between_deposits = AVG(date_diff entre consecutivos). Avg: 8.3 dias', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (LAG dep_date para calcular intervalo)')

# ============================================================
# ADICIONAR fact_redeposits (NOVA TABELA - VERDE)
# ============================================================
# Find last row
last_row = 85
for row in range(85, 130):
    if ws[f'E{row}'].value:
        last_row = row

r = last_row + 2

redeposit_kpis = [
    ('Redeposits', 'fact_redeposits', 'FACT', 'Total depositos por player', 'COUNT de todos depositos confirmados por jogador', 3,
     'COUNT(*) por c_ecr_id WHERE txn_confirmed_success. 154.980 players', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit'),
    ('Redeposits', 'fact_redeposits', 'FACT', 'Redeposit count', 'Depositos apos o FTD (total - 1)', 3,
     'total_deposits - 1. 72.663 redepositors (46.9%)', 'Calculado', 'multibet', 'fact_redeposits'),
    ('Redeposits', 'fact_redeposits', 'FACT', 'Is redepositor D7', 'Flag: 2o deposito em ate 7 dias', 3,
     'CASE WHEN date_diff(ftd_date, second_deposit_date) <= 7 THEN 1. Rate: 39.3%', 'Calculado', 'multibet', 'fact_redeposits'),
    ('Redeposits', 'fact_redeposits', 'FACT', 'Dias ate 2o deposito', 'Dias entre FTD e segundo deposito', 3,
     'date_diff(day, ftd_date, second_deposit_date)', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (ROW_NUMBER rn=1 vs rn=2)'),
    ('Redeposits', 'fact_redeposits', 'FACT', 'Avg redeposit amount', 'Ticket medio dos redepositos', 3,
     'AVG(dep_amount) WHERE dep_rank > 1. Avg: R$ 190.24', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (c_confirmed_amount WHERE rn > 1)'),
    ('Redeposits', 'fact_redeposits', 'FACT', 'Avg dias entre depositos', 'Intervalo medio entre depositos consecutivos', 3,
     'AVG(date_diff entre LAG e atual). Avg: 8.3 dias', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (LAG para calcular gaps)'),
    ('Redeposits', 'fact_redeposits', 'FACT', 'Deposits per month', 'Frequencia mensal de depositos', 3,
     'total_deposits / (dias desde FTD / 30)', 'Calculado', 'multibet', 'fact_redeposits'),
]

for kpi in redeposit_kpis:
    ws[f'B{r}'] = kpi[0]; ws[f'C{r}'] = kpi[1]; ws[f'D{r}'] = kpi[2]
    ws[f'E{r}'] = kpi[3]; ws[f'F{r}'] = kpi[4]; ws[f'G{r}'] = kpi[5]
    ws[f'H{r}'] = kpi[6]; ws[f'I{r}'] = kpi[7]; ws[f'J{r}'] = kpi[8]; ws[f'K{r}'] = kpi[9]
    for col in 'BCDEFGHIJK':
        ws[f'{col}{r}'].font = green_font
    r += 1

wb.save('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
print(f'Excel atualizado! {len(redeposit_kpis)} novas linhas em verde + 8 pendentes resolvidos')

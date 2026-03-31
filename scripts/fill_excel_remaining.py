"""Preenche os 42 KPIs restantes no Excel com fonte/justificativa."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

INPUT = "C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx"
wb = load_workbook(INPUT)
ws = wb['Mapa de KPIs']
updates = 0

for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    kpi = str(row[4].value or '')
    formula = row[7].value
    tabela = str(row[2].value or '')
    dominio = str(row[1].value or '')

    if formula is not None and str(formula).strip() != '':
        continue

    # Marketing e CRM (P6) - BigQuery Smartico
    if 'Marketing e CRM' in dominio:
        if tabela == 'fact_campaigns':
            row[7].value = 'BigQuery Smartico: dm_campaign_stats ou Google Ads API'
            row[8].value = 'BigQuery + Externo'
            row[9].value = 'smartico-bq6 + Google Ads'
            row[10].value = 'Fase 2: dm_campaign_stats / Google Ads API'
        elif tabela == 'fact_crm_communications':
            row[7].value = 'BigQuery Smartico: g_messages (open/click/delivery events)'
            row[8].value = 'BigQuery'
            row[9].value = 'smartico-bq6'
            row[10].value = 'Fase 2: g_messages + g_message_events'
        elif tabela == 'fact_promotions':
            row[7].value = 'BigQuery Smartico: j_bonuses + Athena bonus_ec2'
            row[8].value = 'BigQuery + Athena'
            row[9].value = 'smartico-bq6 + bonus_ec2'
            row[10].value = 'Fase 2: j_bonuses (Smartico) + tbl_ecr_bonus_details (Athena)'
        elif tabela == 'agg_marketing_efficiency':
            row[7].value = 'Calculado: fact_campaigns + fact_gaming_activity_daily'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'Fase 2: fact_campaigns + fact_gaming_activity_daily'
        else:
            continue
        updates += 1

    # Operacoes e Plataforma (P-) - fora do escopo dados
    elif 'Plataforma' in dominio:
        if tabela == 'fact_platform_events':
            row[7].value = 'FORA DO ESCOPO DADOS: metricas de infra (CloudWatch/Datadog)'
            row[8].value = 'Infra/DevOps'
            row[9].value = '-'
            row[10].value = 'CloudWatch, Datadog ou sistema de monitoring (nao Athena/BigQuery)'
        elif tabela == 'fact_support_tickets':
            row[7].value = 'FORA DO ESCOPO DADOS: sistema de suporte (Zendesk/Freshdesk)'
            row[8].value = 'Suporte'
            row[9].value = '-'
            row[10].value = 'Zendesk/Freshdesk API (nao Athena/BigQuery)'
        elif tabela == 'fact_payment_processing':
            if 'PSP' in kpi or 'aprovacao' in kpi.lower():
                row[7].value = 'COUNT(txn_confirmed_success) / COUNT(*) GROUP BY c_processor_name'
                row[8].value = 'Athena'
                row[9].value = 'cashier_ec2'
                row[10].value = 'tbl_cashier_deposit (c_txn_status, c_processor_name)'
            elif 'saque' in kpi.lower() or 'Tempo' in kpi:
                row[7].value = 'AVG(date_diff(c_created_time, c_updated_time)) em horas'
                row[8].value = 'Athena'
                row[9].value = 'cashier_ec2'
                row[10].value = 'tbl_cashier_cashout (c_created_time, c_updated_time, c_txn_status)'
            elif 'Filas' in kpi or 'pendentes' in kpi.lower():
                row[7].value = 'COUNT WHERE c_txn_status IN (pending, processing)'
                row[8].value = 'Athena'
                row[9].value = 'cashier_ec2'
                row[10].value = 'tbl_cashier_cashout (c_txn_status)'
            elif 'Custo' in kpi:
                row[7].value = 'SUM(c_deposit_fee_amount) / SUM(c_deposit_amount) * 100'
                row[8].value = 'Athena'
                row[9].value = 'cashier_ec2'
                row[10].value = 'tbl_cashier_ecr_daily_payment_summary (fee_amount, deposit_amount)'
            elif 'PIX' in kpi:
                row[7].value = 'COUNT_IF(c_option LIKE PIX) / COUNT(*) * 100'
                row[8].value = 'Athena'
                row[9].value = 'cashier_ec2'
                row[10].value = 'tbl_cashier_ecr_daily_payment_summary (c_option)'
            else:
                continue
        elif tabela == 'dim_calendar':
            row[7].value = 'Tabela manual: generate_series de datas com flags (feriado, dia_semana, evento)'
            row[8].value = 'Manual'
            row[9].value = 'multibet'
            row[10].value = 'dim_calendar (criar manualmente com generate_series)'
        else:
            continue
        updates += 1

    # Risco - agg_risk_exposure
    elif tabela == 'agg_risk_exposure':
        if 'Liability' in kpi:
            row[7].value = 'MAX(c_total_stake * TRY_CAST(c_total_odds)) por evento aberto'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bets_info WHERE c_bet_state=O (open bets)'
        elif 'Concentra' in kpi.lower():
            row[7].value = 'SUM(top 10 apostas) / SUM(total) * 100'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bets_info (c_total_stake ORDER BY DESC)'
        elif 'Hedging' in kpi:
            row[7].value = 'PENDENTE: requer sistema de hedging (nao disponivel no Athena)'
            row[8].value = 'N/A'
            row[9].value = '-'
            row[10].value = 'Sistema de trading/hedging (externo)'
        elif 'Perda' in kpi or 'cenario' in kpi.lower():
            row[7].value = 'SUM(c_total_stake * (TRY_CAST(c_total_odds)-1)) WHERE c_bet_state=O'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bets_info (worst case = todas apostas ganham)'
        else:
            continue
        updates += 1

    # Churn - receita reativados
    elif 'Receita' in kpi and 'reativad' in kpi.lower():
        row[7].value = 'Fase 2: SUM(GGR) WHERE player reativado via CRM (requer Smartico)'
        row[8].value = 'BigQuery + Athena'
        row[9].value = 'smartico-bq6 + fund_ec2'
        row[10].value = 'Fase 2: g_messages (CRM) + tbl_real_fund_txn (GGR pos-reativacao)'
        updates += 1

wb.save(INPUT)
print(f'{updates} KPIs preenchidos. Total NaN restante: {42 - updates}')

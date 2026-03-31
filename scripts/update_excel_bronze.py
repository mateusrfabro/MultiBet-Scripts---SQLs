"""
Atualiza o Excel igaming_kpis_v2.xlsx com as fontes Bronze corretas.
Preenche KPIs que estavam NaN com formulas de dados, fonte e tabela origem.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

INPUT = "C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx"
OUTPUT = INPUT  # sobrescreve

wb = load_workbook(INPUT)
ws = wb['Mapa de KPIs']

updates = 0

for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    kpi = str(row[4].value or '')      # E = KPI
    formula_dados = row[7].value       # H = Formula Dados
    tabela = str(row[2].value or '')   # C = Tabela destino

    # Pular se ja tem formula preenchida
    if formula_dados is not None and str(formula_dados).strip() != '':
        continue

    # --- PRODUTO: fact_casino_rounds ---
    if tabela == 'fact_casino_rounds':
        if 'GGR por jogo' in kpi:
            row[7].value = 'Sub-Fund Isolation por c_game_id. SUM(bet-win) GROUP BY game_id, dia'
            row[8].value = 'Athena'
            row[9].value = 'fund_ec2'
            row[10].value = 'tbl_real_fund_txn + sub-funds + tbl_real_fund_txn_type_mst'
        elif 'Hold Rate' in kpi:
            row[7].value = 'GGR / Total Bets * 100 por jogo'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'fact_casino_rounds (ggr, total_bets)'
        elif 'RTP' in kpi:
            row[7].value = 'Total Wins / Total Bets * 100 por slot'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'fact_casino_rounds (total_wins, total_bets)'
        elif 'Rodadas' in kpi:
            row[7].value = 'SUM(c_game_played_count) / COUNT(sessions) por jogo'
            row[8].value = 'Athena'
            row[9].value = 'bireports_ec2'
            row[10].value = 'tbl_ecr_gaming_sessions (c_game_played_count)'
        elif 'Top 20' in kpi:
            row[7].value = 'ORDER BY GGR DESC LIMIT 20. JOIN dim_games_catalog para nome'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'fact_casino_rounds + bronze_games_catalog'
        elif 'provedor' in kpi.lower():
            row[7].value = 'SUM(GGR) GROUP BY c_sub_vendor_id. JOIN games_catalog para nome'
            row[8].value = 'Athena'
            row[9].value = 'fund_ec2'
            row[10].value = 'tbl_real_fund_txn (c_sub_vendor_id) + tbl_vendor_games_mapping_mst'
        else:
            continue
        updates += 1
        print(f'  [casino_rounds] {kpi}')

    # --- PRODUTO: fact_sports_bets ---
    elif tabela == 'fact_sports_bets':
        if 'Turnover' in kpi and 'esporte' not in kpi.lower():
            row[7].value = 'SUM(c_total_stake) WHERE c_transaction_type=M'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bets_info (c_total_stake)'
        elif 'GGR Sports' in kpi:
            row[7].value = 'SUM(c_total_stake) - SUM(c_total_return) WHERE settled'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bets_info (c_total_stake, c_total_return)'
        elif 'Margin' in kpi or 'hold' in kpi.lower():
            row[7].value = 'GGR Sports / Turnover * 100'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'fact_sports_bets (ggr, turnover)'
        elif 'Ticket' in kpi:
            row[7].value = 'AVG(c_total_stake) WHERE c_transaction_type=M'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bets_info (c_total_stake)'
        elif 'pre-live' in kpi.lower() or 'ao vivo' in kpi.lower():
            row[7].value = 'COUNT_IF(c_bet_type=PreLive) / COUNT(*) * 100'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bets_info (c_bet_type)'
        elif 'Top esportes' in kpi:
            row[7].value = 'SUM(c_total_stake) GROUP BY c_sport_type_name ORDER BY DESC'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bet_details (c_sport_type_name)'
        elif 'settled' in kpi.lower() or 'Proje' in kpi:
            row[7].value = 'SUM(c_total_stake * TRY_CAST(c_total_odds AS DOUBLE)) WHERE c_bet_state=O'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bets_info (c_total_stake, c_total_odds, c_bet_closure_time)'
        else:
            continue
        updates += 1
        print(f'  [sports_bets] {kpi}')

    # --- PRODUTO: fact_live_casino ---
    elif tabela == 'fact_live_casino':
        if 'GGR Live' in kpi:
            row[7].value = 'Sub-Fund Isolation WHERE c_game_category LIKE Live%. SUM(bet-win)'
            row[8].value = 'Athena'
            row[9].value = 'fund_ec2'
            row[10].value = 'tbl_real_fund_txn + sub-funds WHERE game_category=Live'
        elif 'Turnover' in kpi:
            row[7].value = 'SUM(bets) por game_id WHERE game_category LIKE Live%'
            row[8].value = 'Athena'
            row[9].value = 'fund_ec2'
            row[10].value = 'tbl_real_fund_txn WHERE game_category=Live'
        elif 'Ocupa' in kpi:
            row[7].value = 'PENDENTE: dados de ocupacao nao disponiveis no Athena'
            row[8].value = 'N/A'
            row[9].value = '-'
            row[10].value = 'Nao disponivel (infra provedor)'
        elif 'Tempo' in kpi and 'sess' in kpi.lower():
            row[7].value = 'AVG(c_session_length_in_sec) / 60 WHERE c_game_category LIKE Live%'
            row[8].value = 'Athena'
            row[9].value = 'bireports_ec2'
            row[10].value = 'tbl_ecr_gaming_sessions (c_session_length_in_sec)'
        else:
            continue
        updates += 1
        print(f'  [live_casino] {kpi}')

    # --- CATALOGOS: dim_games_catalog ---
    elif tabela == 'dim_games_catalog':
        if 'ativos' in kpi.lower():
            row[7].value = 'COUNT(*) WHERE c_status=active vs inactive'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_vendor_games_mapping_mst (c_status)'
        elif 'categoria' in kpi.lower() or 'Mix' in kpi:
            row[7].value = 'COUNT GROUP BY c_game_category_desc / total * 100'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_vendor_games_mapping_mst (c_game_category_desc)'
        elif 'esportes' in kpi.lower() or 'Coverage' in kpi:
            row[7].value = 'COUNT(DISTINCT c_sport_type_name)'
            row[8].value = 'Athena'
            row[9].value = 'vendor_ec2'
            row[10].value = 'tbl_sports_book_bet_details (c_sport_type_name)'
        else:
            continue
        updates += 1
        print(f'  [dim_games] {kpi}')

    # --- AGG: agg_game_performance ---
    elif tabela == 'agg_game_performance':
        if 'rank' in kpi.lower() or 'Receita por jogo' in kpi:
            row[7].value = 'RANK() OVER (ORDER BY GGR DESC) por semana'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'fact_casino_rounds (ggr por game_id)'
        elif 'nicos' in kpi.lower() or 'DAU por jogo' in kpi:
            row[7].value = 'COUNT(DISTINCT c_ecr_id) GROUP BY game_id, dia'
            row[8].value = 'Athena'
            row[9].value = 'fund_ec2'
            row[10].value = 'tbl_real_fund_txn (c_ecr_id, c_game_id)'
        elif 'Concentra' in kpi:
            row[7].value = 'SUM(GGR top 10% jogos) / SUM(GGR total) * 100'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'fact_casino_rounds (GGR por game_id)'
        elif 'estreantes' in kpi.lower() or 'legado' in kpi.lower():
            row[7].value = 'Compare GGR jogos lancados <30d vs >30d via c_updated_time'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'fact_casino_rounds + bronze_games_catalog (c_updated_time)'
        else:
            continue
        updates += 1
        print(f'  [agg_game] {kpi}')

    # --- JACKPOTS ---
    elif tabela == 'fact_jackpots':
        if 'pagos' in kpi.lower():
            row[7].value = 'COUNT(*) WHERE c_txn_type=jackpot_win por mes'
            row[8].value = 'Athena'
            row[9].value = 'fund_ec2'
            row[10].value = 'tbl_real_fund_txn + type_mst (jackpot txn types)'
        elif 'Valor' in kpi and 'jackpot' in kpi.lower():
            row[7].value = 'AVG(jackpot_amount) por mes'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'fact_jackpots (total_amount / count)'
        elif 'Impacto' in kpi:
            row[7].value = 'SUM(jackpot_amount) / GGR * 100 no periodo'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'fact_jackpots + fact_gaming_activity_daily'
        else:
            continue
        updates += 1
        print(f'  [jackpots] {kpi}')

    # --- RISCO: fact_fraud_alerts ---
    elif tabela == 'fact_fraud_alerts':
        if 'Alertas' in kpi:
            row[7].value = 'COUNT(*) WHERE c_ccf_score > threshold por dia'
            row[8].value = 'Athena'
            row[9].value = 'risk_ec2'
            row[10].value = 'tbl_ecr_ccf_score (c_ccf_score)'
        elif 'falso positivo' in kpi.lower():
            row[7].value = 'PENDENTE: requer feedback manual (nao disponivel no Athena)'
            row[8].value = 'N/A'
            row[9].value = '-'
            row[10].value = 'Requer sistema de feedback'
        elif 'bloqueadas' in kpi.lower():
            row[7].value = 'COUNT WHERE c_ecr_status=blocked ou suspended'
            row[8].value = 'Athena'
            row[9].value = 'ecr_ec2'
            row[10].value = 'tbl_ecr (c_ecr_status)'
        elif 'disputa' in kpi.lower():
            row[7].value = 'SUM(c_cb_amount) / 100 WHERE em analise'
            row[8].value = 'Athena'
            row[9].value = 'cashier_ec2'
            row[10].value = 'tbl_cashier_ecr_daily_payment_summary (c_cb_amount)'
        elif 'Multi-account' in kpi:
            row[7].value = 'PENDENTE: requer logica IP/device fingerprint'
            row[8].value = 'Athena'
            row[9].value = 'ecr_ec2'
            row[10].value = 'tbl_ecr_signup_info (c_ip, c_device_finger_print)'
        else:
            continue
        updates += 1
        print(f'  [fraud] {kpi}')

    # --- RISCO: fact_payment_risk ---
    elif tabela == 'fact_payment_risk':
        if 'Chargeback' in kpi:
            row[7].value = 'c_cb_count / (c_deposit_count + c_success_cashout_count) * 100'
            row[8].value = 'Athena'
            row[9].value = 'cashier_ec2'
            row[10].value = 'tbl_cashier_ecr_daily_payment_summary (c_cb_count, c_cb_amount)'
        else:
            continue
        updates += 1
        print(f'  [payment_risk] {kpi}')

    # --- KYC ---
    elif tabela == 'fact_kyc_compliance':
        if 'Taxa' in kpi and 'KYC' in kpi:
            row[7].value = 'COUNT(c_level IN (KYC_1,KYC_2)) / COUNT(*) * 100'
            row[8].value = 'Athena'
            row[9].value = 'ecr_ec2'
            row[10].value = 'tbl_ecr_kyc_level (c_level, c_grace_action_status)'
        elif 'pendentes' in kpi.lower():
            row[7].value = 'COUNT WHERE c_grace_action_status NOT IN (approved)'
            row[8].value = 'Athena'
            row[9].value = 'ecr_ec2'
            row[10].value = 'tbl_ecr_kyc_level (c_grace_action_status)'
        elif 'Rejei' in kpi:
            row[7].value = 'COUNT WHERE c_grace_action_status = rejected por mes'
            row[8].value = 'Athena'
            row[9].value = 'ecr_ec2'
            row[10].value = 'tbl_ecr_kyc_level (c_grace_action_status, c_updated_time)'
        elif 'Tempo' in kpi and 'KYC' in kpi:
            row[7].value = 'PENDENTE: nao ha campo de data envio vs data aprovacao no Athena'
            row[8].value = 'N/A'
            row[9].value = '-'
            row[10].value = 'Campos insuficientes em tbl_ecr_kyc_level'
        else:
            continue
        updates += 1
        print(f'  [kyc] {kpi}')

    # --- RFM Segments ---
    elif tabela == 'agg_player_segments':
        if 'RFM' in kpi:
            row[7].value = 'Recency: days_since_last_bet. Frequency: COUNT bets/30d. Monetary: SUM(GGR) 30d. Score 1-5 por quartil'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'bronze_real_fund_txn + bronze_cashier_deposit (agregado player)'
        elif 'segmento' in kpi.lower() or 'tier' in kpi.lower():
            row[7].value = 'SUM(NGR) GROUP BY RFM_tier (VIP/mid/casual)'
            row[8].value = 'Calculado'
            row[9].value = 'multibet'
            row[10].value = 'agg_player_segments + fact_gaming_activity_daily'
        else:
            continue
        updates += 1
        print(f'  [segments] {kpi}')

wb.save(OUTPUT)
print(f'\n=== {updates} KPIs atualizados no Excel ===')
print(f'Salvo em: {OUTPUT}')

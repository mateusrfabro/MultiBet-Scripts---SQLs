"""
Atualiza o Excel igaming_kpis_v2.xlsx com todas as logicas, colunas H/I/J/K,
e novas tabelas (fct_casino_activity, fct_sports_activity) com fonte verde.
"""
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
ws = wb['Mapa de KPIs']

green_font = Font(color="00B050", bold=False)

def set_cells(row, h=None, i=None, j=None, k=None, green=False):
    for col, val in [('H', h), ('I', i), ('J', j), ('K', k)]:
        if val:
            cell = ws[f'{col}{row}']
            cell.value = val
            if green:
                cell.font = green_font

# ============================================================
# P2: fact_registrations (34-39)
# ============================================================
set_cells(34, 'COUNT(DISTINCT c_ecr_id) GROUP BY dt (dia BRT)', 'Athena', 'bireports_ec2', 'tbl_ecr (c_ecr_id, c_sign_up_time)')
set_cells(35, 'COUNT(DISTINCT c_ecr_id) WHERE ROW_NUMBER()=1 ORDER BY c_created_time. Dia do DEPOSITO. INNER JOIN registrations', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (c_ecr_id, c_created_time, txn_confirmed_success)')
set_cells(36, '(qty_ftds / qty_registrations) * 100. CASE WHEN evita divisao por zero', 'Calculado', 'multibet', 'fact_registrations (qty_ftds, qty_registrations)')
set_cells(37, 'AVG(date_diff(second, c_sign_up_time, ftd_time)) / 3600.0. HORAS decimais. NULL sem FTDs', 'Athena', 'bireports_ec2 + cashier_ec2', 'tbl_ecr + tbl_cashier_deposit')
set_cells(38, '(COUNT(DISTINCT kyc.c_ecr_id) / qty_registrations) * 100. KYC = c_level IN (KYC_1, KYC_2)', 'Athena', 'ecr_ec2', 'tbl_ecr_kyc_level (c_level)')
set_cells(39, 'COUNT_IF(device = Mobile/Desktop/Tablet/Nao Informado). Device via c_channel', 'Athena', 'ecr_ec2', 'tbl_ecr_signup_info (c_channel)')

# ============================================================
# P2: fact_ftd_deposits (40-43)
# ============================================================
set_cells(40, 'AVG(c_confirmed_amount_in_inhouse_ccy / 100.0) WHERE rn=1. Grao: dt x tracker', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (c_confirmed_amount_in_inhouse_ccy)')
set_cells(41, 'PENDENTE: requer campo c_processor_name', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (c_processor_name)')
set_cells(42, 'COUNT_IF por faixa: <50 / 50-500 / >=500. Value Bands Low/Mid/High', 'Calculado', 'multibet', 'fact_ftd_deposits (qty_ftds_below_50, 50_to_500, above_500)')
set_cells(43, 'PENDENTE: requer JOIN com bonus_ec2', 'Athena', 'cashier_ec2 + bonus_ec2', 'tbl_cashier_deposit + bonus tables')

# ============================================================
# P2: fact_gaming_activity_daily (44-52) — Sub-Fund Isolation v2
# ============================================================
set_cells(44, 'Sub-Fund Isolation v2. JOINs: realcash + bonus + type_mst + ecr_flags. test_user=false. Grao: dt x tracker', 'Athena', 'fund_ec2', 'tbl_real_fund_txn + sub-tabelas + tbl_real_fund_txn_type_mst + ecr_flags')
set_cells(45, 'SUM por c_product_id=CASINO. Real+Bonus via sub-fund. op_type DB/CR + c_is_cancel_txn', 'Athena', 'fund_ec2', 'tbl_real_fund_txn WHERE c_product_id=CASINO')
set_cells(46, 'SUM por c_product_id=SPORTS_BOOK. Mesma logica sub-fund', 'Athena', 'fund_ec2', 'tbl_real_fund_txn WHERE c_product_id=SPORTS_BOOK')
set_cells(47, 'CASE WHEN total_bets > 0 THEN (GGR / total_bets) * 100', 'Calculado', 'multibet', 'fact_gaming_activity_daily (ggr, total_bets)')
set_cells(48, 'COUNT(DISTINCT c_ecr_id) por dia+tracker. Gatekeeper: safra + COALESCE(tracker, affiliate)', 'Athena', 'fund_ec2 + bireports_ec2', 'tbl_real_fund_txn + tbl_ecr (Gatekeeper)')
set_cells(49, 'MAX(real_val + bonus_val) WHERE op_type=CR. Maior premio no dia/tracker', 'Athena', 'fund_ec2', 'Sub-tabelas (max single win)')
set_cells(50, 'COUNT/SUM WHERE c_is_cancel_txn=true. Saude tecnica do provedor', 'Athena', 'fund_ec2', 'tbl_real_fund_txn_type_mst (c_is_cancel_txn)')
set_cells(51, 'Sub-Fund: Bonus Bets - Bonus Wins via crp+wrp+rrp. Resultado: -R$ 551k (casa ganhou)', 'Athena', 'fund_ec2', 'tbl_bonus_sub_fund_txn (crp + wrp + rrp)')
set_cells(52, 'NGR = GGR Real (sem bonus). R$ 15.96M. Validado AWS Console + Mauro (Nov/25: 0.000%)', 'Calculado', 'multibet', 'fact_gaming_activity_daily (ngr)')

# ============================================================
# P2: fact_attribution (53-58)
# ============================================================
set_cells(53, 'marketing_spend / qty_ftds. Spend proporcional por FTDs do tracker Google', 'Calculado', 'multibet', 'fact_attribution + vw_attribution_metrics')
set_cells(54, 'marketing_spend / qty_registrations', 'Calculado', 'multibet', 'fact_attribution + vw_attribution_metrics')
set_cells(55, 'ggr / marketing_spend. Fonte GGR: gaming. Fonte Spend: Google Sheets', 'Calculado', 'multibet', 'vw_attribution_metrics (roas)')
set_cells(56, 'GGR por tracker via gaming. JOIN dim_marketing_mapping para source', 'Athena + Super Nova', 'fund_ec2 + multibet', 'fact_gaming_activity_daily + dim_marketing_mapping')
set_cells(57, 'PENDENTE: requer tabela de contratos/comissoes afiliados', 'A definir', '-', 'Tabela contratos (a criar)')
set_cells(58, 'PENDENTE: LTV via agg_cohort / CAC via vw_attribution_metrics', 'Calculado', 'multibet', 'agg_cohort_acquisition + vw_attribution_metrics')

# ============================================================
# P2: dim_marketing_mapping (59-62)
# ============================================================
set_cells(59, 'COALESCE(tracker, affiliate, sem_tracker). 3.241 trackers com source padronizado (IDs oficiais)', 'Athena + Manual', 'bireports_ec2 + ecr_ec2', 'tbl_ecr + tbl_ecr_banner (URLs forense)')
set_cells(60, 'High (Official/GCLID/FBCLID), Medium (campaign_id/AFP), Low (generico)', 'Manual', 'multibet', 'dim_marketing_mapping (confidence)')
set_cells(61, 'Texto: URLs com gclid=, fbclid=, utm_source=, afp=. Auditoria 19/03', 'Manual', 'multibet', 'dim_marketing_mapping (mapping_logic)')
set_cells(62, '3.241 trackers mapeados. Cobertura por GGR a validar', 'Calculado', 'multibet', 'dim_marketing_mapping vs fact_attribution')

# ============================================================
# P2: dim_acquisition_channel (63-64)
# ============================================================
set_cells(63, 'VIEW tiering: Paid Media, Partnerships, Direct/Organic', 'Super Nova DB', 'multibet', 'vw_acquisition_channel')
set_cells(64, 'FTD Rate e ROAS por canal/tier', 'Calculado', 'multibet', 'vw_acquisition_channel')

# ============================================================
# P2: agg_cohort_acquisition (65-68)
# ============================================================
set_cells(65, 'AVG(ggr_d0, ggr_d7, ggr_d30) por month_of_ftd x source. Player-level. 154.960 players', 'Athena', 'fund_ec2 + cashier_ec2', 'tbl_real_fund_txn + tbl_cashier_deposit + dim_marketing_mapping')
set_cells(66, 'is_2nd_depositor = ROW_NUMBER rn=2. Global: 46.9%', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit (2nd deposit detection)')
set_cells(67, 'monthly_spend / payback via vw_cohort_roi', 'Calculado', 'multibet', 'vw_cohort_roi')
set_cells(68, 'SUM(ggr_d30) por cohort. Total acumulado desde FTD', 'Calculado', 'multibet', 'agg_cohort_acquisition + vw_cohort_roi')

# ============================================================
# P3: fact_player_activity (69-74)
# ============================================================
set_cells(69, 'COUNT(DISTINCT c_ecr_id) janela deslizante. DAU 4.249, MAU 36.435', 'Athena', 'fund_ec2 + bireports_ec2', 'tbl_real_fund_txn + tbl_ecr')
set_cells(70, '= DAU. COUNT DISTINCT jogadores com bets no dia', 'Athena', 'fund_ec2', 'tbl_real_fund_txn')
set_cells(71, 'DAU / MAU * 100. Stickiness 11.7%', 'Calculado', 'multibet', 'fact_player_activity (dau, mau)')
set_cells(72, 'total_bets / dau = avg_bets_per_player', 'Calculado', 'multibet', 'fact_player_activity')
set_cells(73, 'PENDENTE: requer tbl_real_fund_session (timestamps inicio/fim)', 'Athena', 'fund_ec2', 'tbl_real_fund_session (se existir)')
set_cells(74, 'total_ggr / dau = ggr_per_dau', 'Calculado', 'multibet', 'fact_player_activity')

# ============================================================
# P3: fact_redeposits (75-78)
# ============================================================
set_cells(75, 'is_2nd_depositor na agg_cohort. Global 46.9%. Google 53%, Meta 49%', 'Athena + Super Nova', 'cashier_ec2 + multibet', 'agg_cohort_acquisition')
set_cells(76, 'PENDENTE: COUNT todos depositos por player', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit')
set_cells(77, 'PENDENTE: AVG(amount) depositos pos-FTD', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit')
set_cells(78, 'PENDENTE: AVG(date_diff entre depositos consecutivos)', 'Athena', 'cashier_ec2', 'tbl_cashier_deposit')

# ============================================================
# P3: fact_churn_events (79-82)
# ============================================================
set_cells(79, 'is_churned=1 se days_since_last_active > 30. Taxa 71.1%', 'Super Nova DB', 'multibet', 'fact_player_engagement_daily')
set_cells(80, 'days_since_last_active por player. D30/D60/D90 via WHERE', 'Super Nova DB', 'multibet', 'fact_player_engagement_daily')
set_cells(81, 'PENDENTE: requer CRM Smartico (campanhas reativacao)', 'BigQuery', 'smartico-bq6', 'Views CRM + fact_player_engagement_daily')
set_cells(82, 'PENDENTE: spend reativacao / count reativados', 'BigQuery + Manual', 'smartico + multibet', 'CRM + fact_player_engagement_daily')

# ============================================================
# NOVAS TABELAS: fct_casino_activity + fct_sports_activity (VERDE)
# ============================================================
last_row = 85
for row in range(85, 120):
    if ws[f'E{row}'].value is None:
        # Check if previous also empty
        if row > 85 and ws[f'E{row-1}'].value is None:
            last_row = row - 1
            break
        last_row = row

r = last_row + 1

# Clear old casino/sports rows if they exist
for check_row in range(85, 120):
    val = str(ws[f'C{check_row}'].value or '')
    if 'fct_casino' in val or 'fct_sports' in val:
        for col in 'ABCDEFGHIJK':
            ws[f'{col}{check_row}'] = None

r = last_row + 1

casino_kpis = [
    ('Casino (Sub-Fund)', 'fct_casino_activity', 'FACT', 'Casino Real Bet', 'Apostas em dinheiro real por dia', 2,
     'Sub-Fund Isolation: SUM(realcash+drp) WHERE op_type=DB. Validado AWS Console 100%', 'Athena', 'fund_ec2', 'tbl_realcash_sub_fund_txn + tbl_bonus_sub_fund_txn + tbl_real_fund_txn_type_mst'),
    ('Casino (Sub-Fund)', 'fct_casino_activity', 'FACT', 'Casino Real Win', 'Ganhos em dinheiro real por dia', 2,
     'Sub-Fund: SUM(realcash+drp) WHERE op_type=CR', 'Athena', 'fund_ec2', 'tbl_realcash_sub_fund_txn + tbl_bonus_sub_fund_txn'),
    ('Casino (Sub-Fund)', 'fct_casino_activity', 'FACT', 'Casino Real GGR', 'Real Bet - Real Win', 2,
     'Nov/25: diff R$ 8 em R$ 282k vs Mauro (0.003%). GGR Real R$ 23.6M', 'Calculado', 'multibet', 'fct_casino_activity'),
    ('Casino (Sub-Fund)', 'fct_casino_activity', 'FACT', 'Casino Bonus Bet', 'Apostas com bonus (crp+wrp+rrp)', 2,
     'Sub-Fund: SUM(crp+wrp+rrp) WHERE op_type=DB', 'Athena', 'fund_ec2', 'tbl_bonus_sub_fund_txn'),
    ('Casino (Sub-Fund)', 'fct_casino_activity', 'FACT', 'Casino Bonus Win', 'Ganhos com bonus', 2,
     'Sub-Fund: SUM(crp+wrp+rrp) WHERE op_type=CR', 'Athena', 'fund_ec2', 'tbl_bonus_sub_fund_txn'),
    ('Casino (Sub-Fund)', 'fct_casino_activity', 'FACT', 'Casino Total GGR', 'GGR total (Real+Bonus)', 2,
     'Total Bet - Total Win. GGR R$ 20.5M. Filtros: test_user=false, gaming=Y, CASINO', 'Calculado', 'multibet', 'fct_casino_activity'),
]

for kpi in casino_kpis:
    ws[f'B{r}'] = kpi[0]; ws[f'C{r}'] = kpi[1]; ws[f'D{r}'] = kpi[2]
    ws[f'E{r}'] = kpi[3]; ws[f'F{r}'] = kpi[4]; ws[f'G{r}'] = kpi[5]
    ws[f'H{r}'] = kpi[6]; ws[f'I{r}'] = kpi[7]; ws[f'J{r}'] = kpi[8]; ws[f'K{r}'] = kpi[9]
    # Apply green font to all cells in this row
    for col in 'BCDEFGHIJK':
        ws[f'{col}{r}'].font = green_font
    r += 1

sports_kpis = [
    ('Sports (Sub-Fund)', 'fct_sports_activity', 'FACT', 'Sports Real Bet', 'Apostas esportivas em dinheiro real', 2,
     'Sub-Fund Isolation: mesma logica casino, c_product_id=SPORTS_BOOK', 'Athena', 'fund_ec2', 'tbl_realcash_sub_fund_txn + tbl_real_fund_txn'),
    ('Sports (Sub-Fund)', 'fct_sports_activity', 'FACT', 'Sports Real Win', 'Ganhos esportivos em dinheiro real', 2,
     'Sub-Fund: SUM(realcash+drp) WHERE op_type=CR', 'Athena', 'fund_ec2', 'tbl_realcash_sub_fund_txn'),
    ('Sports (Sub-Fund)', 'fct_sports_activity', 'FACT', 'Sports Real GGR', 'GGR real esportivo', 2,
     'Real Bet - Real Win. Validado: -0.03% diff vs Mauro', 'Calculado', 'multibet', 'fct_sports_activity'),
    ('Sports (Sub-Fund)', 'fct_sports_activity', 'FACT', 'Sports Bonus Bet', 'Apostas esportivas com bonus', 2,
     'Sub-Fund: SUM(crp+wrp+rrp) WHERE op_type=DB AND SPORTS_BOOK', 'Athena', 'fund_ec2', 'tbl_bonus_sub_fund_txn'),
    ('Sports (Sub-Fund)', 'fct_sports_activity', 'FACT', 'Sports Total GGR', 'GGR total esportivo', 2,
     'Total Bet - Total Win. GGR R$ 8.2M', 'Calculado', 'multibet', 'fct_sports_activity'),
]

for kpi in sports_kpis:
    ws[f'B{r}'] = kpi[0]; ws[f'C{r}'] = kpi[1]; ws[f'D{r}'] = kpi[2]
    ws[f'E{r}'] = kpi[3]; ws[f'F{r}'] = kpi[4]; ws[f'G{r}'] = kpi[5]
    ws[f'H{r}'] = kpi[6]; ws[f'I{r}'] = kpi[7]; ws[f'J{r}'] = kpi[8]; ws[f'K{r}'] = kpi[9]
    for col in 'BCDEFGHIJK':
        ws[f'{col}{r}'].font = green_font
    r += 1

wb.save('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
print(f'Excel atualizado! {len(casino_kpis) + len(sports_kpis)} novas linhas em verde')
print('Colunas H/I/J/K atualizadas para linhas 34-82')

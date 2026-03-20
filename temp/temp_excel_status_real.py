import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import openpyxl

wb = openpyxl.load_workbook('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
ws = wb.active

# Atualizar linhas 67-72 (fact_player_activity) com status REAL
# Essas serao implementadas pela fact_player_activity (em execucao agora)
updates = {
    67: {
        "H": "EM EXECUCAO: COUNT(DISTINCT c_ecr_id) com janela deslizante: DAU=dia, WAU=7d, MAU=30d",
        "I": "Athena (Pragmatic Solutions)",
        "J": "fund_ec2 + bireports_ec2",
        "K": "fact_player_activity (dau, wau, mau por dia). Fonte: tbl_real_fund_txn (jogadores com bets no periodo)"
    },
    68: {
        "H": "EM EXECUCAO: = DAU (COUNT DISTINCT jogadores com pelo menos 1 bet no dia)",
        "I": "Athena (Pragmatic Solutions)",
        "J": "fund_ec2 + bireports_ec2",
        "K": "fact_player_activity (dau)"
    },
    69: {
        "H": "EM EXECUCAO: DAU / MAU * 100. Stickiness > 20% = saudavel",
        "I": "Super Nova DB (calculado)",
        "J": "multibet",
        "K": "fact_player_activity (stickiness_pct)"
    },
    70: {
        "H": "EM EXECUCAO: total_bets / dau = avg_bets_per_player (proxy de sessoes/intensidade)",
        "I": "Athena + Super Nova DB",
        "J": "fund_ec2 + multibet",
        "K": "fact_player_activity (avg_bets_per_player)"
    },
    71: {
        "H": "PENDENTE - requer tbl_real_fund_session (timestamps inicio/fim de sessao)",
        "I": "Athena",
        "J": "fund_ec2",
        "K": "tbl_real_fund_session - nao implementada ainda"
    },
    72: {
        "H": "EM EXECUCAO: total_ggr / dau = ggr_per_dau (receita por jogador ativo)",
        "I": "Athena + Super Nova DB",
        "J": "fund_ec2 + multibet",
        "K": "fact_player_activity (ggr_per_dau)"
    },
}

for row, data in updates.items():
    ws.cell(row=row, column=8, value=data["H"])
    ws.cell(row=row, column=9, value=data["I"])
    ws.cell(row=row, column=10, value=data["J"])
    ws.cell(row=row, column=11, value=data["K"])

wb.save('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
print("Excel atualizado com status real das tabelas!")

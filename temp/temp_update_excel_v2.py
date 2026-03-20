import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import openpyxl

wb = openpyxl.load_workbook('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
ws = wb.active

# Padronizar TODAS as linhas com formato consistente na coluna H:
# SQL: <formula>  |  Fonte: <database.tabela>  |  Filtros: <condicoes>
# Esse padrao facilita para quem ler entender a logica de cara

# Mapear todas as linhas que preenchemos
updates = {
    # fact_registrations (34-39)
    34: "SQL: COUNT(DISTINCT c_ecr_id) GROUP BY dt  |  Fonte: bireports_ec2.tbl_ecr (c_sign_up_time)  |  Filtros: c_registration_status = ecr_regn_completed, safra >= 2025-10-01",
    35: "SQL: COUNT(DISTINCT c_ecr_id) WHERE ROW_NUMBER()=1, agrupado por dia do DEPOSITO  |  Fonte: cashier_ec2.tbl_cashier_deposit  |  Filtros: c_txn_status = txn_confirmed_success, INNER JOIN registrations (Gatekeeper)",
    36: "SQL: (qty_ftds / qty_registrations) * 100  |  Fonte: multibet.fact_registrations (calculado)  |  Filtros: CASE WHEN regs > 0 para evitar divisao por zero",
    37: "SQL: AVG(date_diff('second', c_sign_up_time, ftd_time)) / 3600.0  |  Fonte: bireports_ec2.tbl_ecr + cashier_ec2.tbl_cashier_deposit  |  Filtros: NULL quando nao ha FTDs no dia",
    38: "SQL: COUNT(DISTINCT kyc.c_ecr_id) / qty_registrations * 100  |  Fonte: ecr_ec2.tbl_ecr_kyc_level  |  Filtros: c_level IN ('KYC_1','KYC_2'). NAO usar tbl_ecr_aml_flags",
    39: "SQL: COUNT_IF(device = 'Mobile/Desktop/Tablet/Nao Informado')  |  Fonte: ecr_ec2.tbl_ecr_signup_info (c_channel)  |  Filtros: web+WEB = Desktop, vazio = Nao Informado",
    # fact_ftd_deposits (40-43)
    40: "SQL: AVG(c_confirmed_amount_in_inhouse_ccy / 100.0) via DECIMAL(18,2)  |  Fonte: cashier_ec2.tbl_cashier_deposit  |  Filtros: ROW_NUMBER()=1 (FTD), Gatekeeper Pattern, COALESCE(tracker, affiliate)",
    41: "PENDENTE  |  Fonte: cashier_ec2.tbl_cashier_deposit (c_processor_name)  |  Filtros: agrupar por metodo de pagamento",
    42: "SQL: COUNT_IF(ftd_amount < 50 / 50-500 / >= 500)  |  Fonte: cashier_ec2.tbl_cashier_deposit  |  Filtros: Value Bands Low/Mid/High por dia x tracker",
    43: "PENDENTE  |  Fonte: cashier_ec2.tbl_cashier_deposit (c_bonus_code) + bonus_ec2  |  Filtros: JOIN com bonus para identificar FTDs com bonus",
    # fact_gaming_activity_daily (44-50)
    44: "SQL: SUM(Bets 27,28,59) - SUM(Wins 45,80,72,112) GROUP BY dt, tracker  |  Fonte: fund_ec2.tbl_real_fund_txn  |  Filtros: c_txn_status = SUCCESS, DECIMAL(18,2), Gatekeeper + COALESCE(tracker, affiliate)",
    45: "SQL: SUM(c_txn_type IN (27,28)) - SUM(c_txn_type IN (45,80,72))  |  Fonte: fund_ec2.tbl_real_fund_txn  |  Filtros: apenas tipos casino",
    46: "SQL: SUM(c_txn_type=59) - SUM(c_txn_type=112)  |  Fonte: fund_ec2.tbl_real_fund_txn  |  Filtros: apenas tipos sportsbook",
    47: "SQL: CASE WHEN SUM(bets) > 0 THEN (GGR / SUM(bets)) * 100  |  Fonte: multibet.fact_gaming_activity_daily (calculado)  |  Filtros: evita divisao por zero",
    48: "SQL: COUNT(DISTINCT c_ecr_id) GROUP BY dt, tracker  |  Fonte: fund_ec2.tbl_real_fund_txn + bireports_ec2.tbl_ecr  |  Filtros: Gatekeeper Pattern + COALESCE(tracker, affiliate)",
    49: "SQL: MAX(casino_win + sports_win) por dia x tracker  |  Fonte: fund_ec2.tbl_real_fund_txn  |  Filtros: contexto para GGR negativo (premio isolado vs tendencia)",
    50: "SQL: SUM(CASE c_txn_type=72 THEN 1) + SUM(rollback_amount)  |  Fonte: fund_ec2.tbl_real_fund_txn  |  Filtros: ~8% dos wins sao rollbacks, monitora estabilidade provedores",
    # fact_attribution (51-56)
    51: "SQL: marketing_spend / qty_ftds via vw_attribution_metrics  |  Fonte: Google Sheets (spend) + cashier_ec2 (FTDs)  |  Filtros: Spend proporcional por FTDs dos trackers google_ads via dim_marketing_mapping",
    52: "SQL: marketing_spend / qty_registrations via vw_attribution_metrics  |  Fonte: Google Sheets + bireports_ec2  |  Filtros: mesma logica de distribuicao proporcional",
    53: "SQL: ggr / marketing_spend via vw_attribution_metrics  |  Fonte: fact_gaming_activity_daily (GGR) + Google Sheets (spend)  |  Filtros: por source (google_ads, meta_ads, etc) via dim_marketing_mapping",
    54: "SQL: SUM(ggr) GROUP BY c_tracker_id, incluindo ggr_casino + ggr_sports  |  Fonte: fund_ec2 + bireports_ec2  |  Filtros: Gatekeeper + COALESCE(tracker, affiliate). 297657 = Google Ads (confirmado forense gclid)",
    55: "PENDENTE  |  Fonte: tabela de contratos afiliados (externa)  |  Filtros: requer % comissao por afiliado",
    56: "PENDENTE  |  Fonte: agg_cohort_acquisition (LTV) + vw_attribution_metrics (CAC)  |  Filtros: LTV D30 implementado, CAC implementado",
}

# dim_marketing_mapping (57-60)
for row in range(57, 61):
    e = ws.cell(row=row, column=5).value or ''
    if 'tracker' in e.lower() and 'source' in e.lower():
        updates[row] = "SQL: COALESCE(NULLIF(TRIM(c_tracker_id),''), CAST(c_affiliate_id AS VARCHAR), 'sem_tracker')  |  Fonte: bireports_ec2.tbl_ecr (auditoria forense URLs: gclid/fbclid/utm)  |  Filtros: 20 trackers mapeados, 445431+297657=Google Ads, 464673/467185/53194=Meta"
    elif 'Confidence' in e:
        updates[row] = "Valores: High (GCLID/FBCLID direto na URL), Medium (UTM/AFP), Low (heuristica)  |  Fonte: dim_marketing_mapping.confidence  |  Filtros: baseado em evidencia da URL do jogador no momento do registro"
    elif 'Mapping' in e and 'logic' in e.lower():
        updates[row] = "Texto descritivo por tracker: ex '445431: URLs com gclid= e gad_source=1'  |  Fonte: dim_marketing_mapping.mapping_logic  |  Filtros: documentacao forense para auditoria"
    elif 'Cobertura' in e:
        updates[row] = "ACAO NECESSARIA: 20/48.564 trackers mapeados. R$16.5M em orphans  |  Fonte: dim_marketing_mapping vs fact_attribution  |  Filtros: Castrin/Marketing devem fornecer DE-PARA completo"

# agg_cohort_acquisition (rows ~63-66)
for row in range(61, 70):
    c = ws.cell(row=row, column=3).value or ''
    e = ws.cell(row=row, column=5).value or ''
    if 'agg_cohort' not in c:
        continue
    if 'LTV' in e and 'cohort' in e:
        updates[row] = "SQL: AVG(ggr_d30) por month_of_ftd x source. Janelas: date_diff do ftd_time  |  Fonte: fund_ec2.tbl_real_fund_txn (GGR D0/D7/D30) + cashier_ec2 (FTD)  |  Filtros: Gatekeeper + COALESCE. Google Ads Fev/26: ROI 110.5% no D30"
    elif '2nd deposit' in e.lower():
        updates[row] = "SQL: is_2nd_depositor = CASE WHEN ROW_NUMBER()=2 EXISTS THEN 1 ELSE 0  |  Fonte: cashier_ec2.tbl_cashier_deposit  |  Filtros: qualquer data apos FTD (maturacao safra). Global: 46.9%"
    elif 'Payback' in e:
        updates[row] = "SQL: monthly_spend / SUM(ggr_d30) via vw_cohort_roi  |  Fonte: fact_attribution (spend) + agg_cohort (GGR D30)  |  Filtros: JOIN por month_of_ftd x source via dim_marketing_mapping"
    elif 'GGR' in e and 'cohort' in e.lower():
        updates[row] = "SQL: SUM(ggr_d30) por month_of_ftd x source via vw_cohort_roi  |  Fonte: fund_ec2 + cashier_ec2 + bireports_ec2  |  Filtros: Gatekeeper. Google Ads Fev/26: R$1.94M GGR D30"

for row, text in updates.items():
    ws.cell(row=row, column=8, value=text)

wb.save('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
print(f"Excel atualizado! {len(updates)} linhas com padrao consistente: SQL | Fonte | Filtros")

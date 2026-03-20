import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import openpyxl

wb = openpyxl.load_workbook('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
ws = wb.active

# Encontrar linhas fact_player_activity e fact_redeposits
for row in range(67, 80):
    c = ws.cell(row=row, column=3).value or ''
    e = ws.cell(row=row, column=5).value or ''
    if c:
        print(f"Row {row}: C={c[:30]} | E={e[:50]}")

# Atualizar fact_player_activity (67-72) - implementadas via fact_player_engagement_daily
updates = {
    67: {
        "H": "COUNT(DISTINCT c_ecr_id) WHERE days_since_last_active = 0 (DAU), <= 7 (WAU), <= 30 (MAU)",
        "I": "Super Nova DB",
        "J": "multibet",
        "K": "fact_player_engagement_daily (last_active_date, days_since_last_active)"
    },
    68: {
        "H": "COUNT(DISTINCT c_ecr_id) WHERE last_active_date = dt",
        "I": "Athena + Super Nova DB",
        "J": "fund_ec2 + bireports_ec2",
        "K": "fact_player_engagement_daily (total_active_days por player)"
    },
    69: {
        "H": "DAU / MAU * 100. Stickiness > 20% = saudavel",
        "I": "Super Nova DB (calculado)",
        "J": "multibet",
        "K": "fact_player_engagement_daily (derivado de DAU e MAU)"
    },
    70: {
        "H": "total_bets_count / total_active_days = avg_bets_per_day (proxy de sessoes)",
        "I": "Athena + Super Nova DB",
        "J": "fund_ec2 + multibet",
        "K": "fact_player_engagement_daily (total_bets_count, total_active_days)"
    },
    71: {
        "H": "PENDENTE - requer dados de sessao (tbl_real_fund_session em fund_ec2)",
        "I": "Athena",
        "J": "fund_ec2",
        "K": "tbl_real_fund_session (c_session_id, timestamps inicio/fim)"
    },
    72: {
        "H": "total_ggr / total_active_days por player, ou SUM(ggr) / COUNT(DISTINCT active_players)",
        "I": "Super Nova DB",
        "J": "multibet",
        "K": "fact_player_engagement_daily (total_ggr, total_active_days)"
    },
}

for row, data in updates.items():
    ws.cell(row=row, column=8, value=data["H"])
    ws.cell(row=row, column=9, value=data["I"])
    ws.cell(row=row, column=10, value=data["J"])
    ws.cell(row=row, column=11, value=data["K"])

# fact_redeposits (73-76) - parcialmente implementadas
updates2 = {
    73: {
        "H": "is_2nd_depositor flag na agg_cohort_acquisition. Global: 46.9%. Por source: Google 53%, Meta 49%",
        "I": "Athena + Super Nova DB",
        "J": "cashier_ec2 + multibet",
        "K": "agg_cohort_acquisition (is_2nd_depositor = ROW_NUMBER rn=2)"
    },
    74: {
        "H": "PENDENTE - requer contagem de todos depositos por player (nao so FTD e 2nd)",
        "I": "Athena",
        "J": "cashier_ec2",
        "K": "tbl_cashier_deposit (COUNT por c_ecr_id WHERE txn_confirmed_success)"
    },
    75: {
        "H": "PENDENTE - AVG(amount) de depositos posteriores ao FTD",
        "I": "Athena",
        "J": "cashier_ec2",
        "K": "tbl_cashier_deposit (c_confirmed_amount_in_inhouse_ccy WHERE rn > 1)"
    },
    76: {
        "H": "PENDENTE - AVG(date_diff entre depositos consecutivos) por player",
        "I": "Athena",
        "J": "cashier_ec2",
        "K": "tbl_cashier_deposit (LAG c_created_time para calcular intervalo)"
    },
}

for row, data in updates2.items():
    ws.cell(row=row, column=8, value=data["H"])
    ws.cell(row=row, column=9, value=data["I"])
    ws.cell(row=row, column=10, value=data["J"])
    ws.cell(row=row, column=11, value=data["K"])

wb.save('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
print(f"Excel atualizado! Linhas 67-76 preenchidas (engagement + redeposits)")

import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import openpyxl

wb = openpyxl.load_workbook('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
ws = wb.active

# Encontrar linhas dim_acquisition_channel
for row in range(61, 66):
    c = ws.cell(row=row, column=3).value or ''
    e = ws.cell(row=row, column=5).value or ''
    if 'dim_acquisition' in c:
        if 'Mix' in e:
            ws.cell(row=row, column=8, value="SUM(qty_registrations/ftds/ggr) GROUP BY channel_tier")
            ws.cell(row=row, column=9, value="Super Nova DB (View)")
            ws.cell(row=row, column=10, value="multibet")
            ws.cell(row=row, column=11, value="vw_acquisition_channel: Direct/Organic (organic), Paid Media (google+meta+tiktok+ig), Partnerships (influencers+portais+affiliates)")
        elif 'Qualidade' in e:
            ws.cell(row=row, column=8, value="FTD Rate + ROAS por channel_tier via vw_acquisition_channel")
            ws.cell(row=row, column=9, value="Super Nova DB (View)")
            ws.cell(row=row, column=10, value="multibet")
            ws.cell(row=row, column=11, value="vw_acquisition_channel: ftd_rate = ftds/regs*100, roas = ggr/spend. Paid Media: 45.5% GGR, Partnerships: 6.7%, Unmapped: 47.7%")

wb.save('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
print("Excel atualizado com dim_acquisition_channel!")

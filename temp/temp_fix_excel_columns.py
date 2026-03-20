import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import openpyxl

wb = openpyxl.load_workbook('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
ws = wb.active

# Para cada linha que tem "|" na coluna H, separar em H/I/J/K
count = 0
for row in range(3, ws.max_row + 1):
    h = ws.cell(row=row, column=8).value
    if not h or '|' not in str(h):
        continue

    parts = str(h).split('|')

    # H = Formula/SQL (primeira parte, remover prefixo "SQL: ")
    formula = parts[0].strip()
    if formula.startswith('SQL: '):
        formula = formula[5:]
    elif formula.startswith('SQL:'):
        formula = formula[4:]
    ws.cell(row=row, column=8, value=formula.strip())

    # I = Fonte (segunda parte, remover prefixo "Fonte: ")
    if len(parts) > 1:
        fonte = parts[1].strip()
        if fonte.startswith('Fonte: '):
            fonte = fonte[7:]
        elif fonte.startswith('Fonte:'):
            fonte = fonte[6:]
        ws.cell(row=row, column=9, value=fonte.strip())

    # Extrair database e tabela da fonte
    if len(parts) > 1:
        fonte_text = parts[1].strip()
        # Tentar extrair database (antes do .)
        if 'ec2' in fonte_text or 'multibet' in fonte_text:
            # Database = parte antes do . ou entre parenteses
            db_parts = []
            for word in fonte_text.replace('Fonte: ', '').split():
                if '_ec2' in word or 'multibet' in word:
                    db_parts.append(word.split('.')[0].strip('(').strip(')'))
            if db_parts:
                ws.cell(row=row, column=10, value=' + '.join(set(db_parts)))

    # K = Filtros/Tabela (terceira parte, remover prefixo "Filtros: ")
    if len(parts) > 2:
        filtros = parts[2].strip()
        if filtros.startswith('Filtros: '):
            filtros = filtros[9:]
        elif filtros.startswith('Filtros:'):
            filtros = filtros[8:]
        ws.cell(row=row, column=11, value=filtros.strip())

    count += 1

wb.save('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
print(f"Excel corrigido! {count} linhas separadas em colunas H/I/J/K")

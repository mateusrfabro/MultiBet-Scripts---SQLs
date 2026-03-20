import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import openpyxl

wb = openpyxl.load_workbook('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
ws = wb.active

# Find agg_cohort_acquisition rows
cohort_rows = []
for row in range(60, 70):
    c = ws.cell(row=row, column=3).value or ''
    e = ws.cell(row=row, column=5).value or ''
    if 'agg_cohort' in c:
        cohort_rows.append((row, e))

for row, e in cohort_rows:
    print(f"Row {row}: {e}")

# Update each row
updates = {}
for row, e in cohort_rows:
    if 'LTV' in e and 'cohort' in e.lower():
        updates[row] = {
            "H": "IMPLEMENTADA: AVG(ggr_d30) por month_of_ftd x source. Janelas D0/D7/D30 calculadas via date_diff do ftd_time. Fonte: fund_ec2.tbl_real_fund_txn (Bets 27,28,59 - Wins 45,80,72,112). Gatekeeper Pattern + COALESCE(tracker, affiliate)",
            "I": "Athena (Pragmatic Solutions) + Super Nova DB",
            "J": "fund_ec2 + cashier_ec2 + bireports_ec2 + multibet",
            "K": "agg_cohort_acquisition (ggr_d0, ggr_d7, ggr_d30) + vw_cohort_roi (agregado)",
        }
    elif '2nd deposit' in e.lower():
        updates[row] = {
            "H": "IMPLEMENTADA: is_2nd_depositor flag (0/1) por player. ROW_NUMBER rn=2 em cashier_ec2.tbl_cashier_deposit. Taxa = SUM(flag) / COUNT(*) * 100. Resultado: 46.9% global",
            "I": "Athena (Pragmatic Solutions)",
            "J": "cashier_ec2",
            "K": "tbl_cashier_deposit (ROW_NUMBER rn=2 para 2nd deposit detection)",
        }
    elif 'Payback' in e:
        updates[row] = {
            "H": "IMPLEMENTADA via vw_cohort_roi: payback_ratio = monthly_spend / SUM(ggr_d30). Spend via dim_marketing_mapping como ponte. Google Ads Out/25: Spend R$5.9M vs GGR D30 R$384k = payback 15.4x (precisa 15 meses D30 para pagar)",
            "I": "Super Nova DB (calculado)",
            "J": "multibet",
            "K": "vw_cohort_roi (payback_ratio) = fact_attribution.spend / agg_cohort.ggr_d30",
        }
    elif 'GGR' in e and 'cohort' in e.lower():
        updates[row] = {
            "H": "IMPLEMENTADA: SUM(ggr_d30) por month_of_ftd x source. Resultados: Multi-channel Fev/26 = R$1.94M, Google Out/25 = R$384k, Meta Jan/26 = R$239k. GGR inclui Casino+Sports com rollbacks",
            "I": "Athena (Pragmatic Solutions) + Super Nova DB",
            "J": "fund_ec2 + multibet",
            "K": "agg_cohort_acquisition (ggr_d30 acumulado) + vw_cohort_roi",
        }

for row, data in updates.items():
    ws.cell(row=row, column=8, value=data["H"])
    ws.cell(row=row, column=9, value=data["I"])
    ws.cell(row=row, column=10, value=data["J"])
    ws.cell(row=row, column=11, value=data["K"])

wb.save('C:/Users/NITRO/Downloads/igaming_kpis_v2.xlsx')
print(f"Excel atualizado! {len(updates)} linhas de agg_cohort_acquisition preenchidas")
